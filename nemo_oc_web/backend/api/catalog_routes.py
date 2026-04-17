"""
Endpoints para carga de catálogos Excel y consulta de estadísticas.
"""
import io
import sys
import shutil
from pathlib import Path

from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse

_nemo_oc_dir = Path(__file__).parent.parent.parent.parent / "nemo_oc"
if str(_nemo_oc_dir) not in sys.path:
    sys.path.insert(0, str(_nemo_oc_dir))

from app.config import (
    CORREOS_FILENAME,
    LICITACIONES_FILENAME,
    REDSALUD_HOMO_FILENAME,
    CARTERA_FILENAME,
    HOMOLOGACION_FILENAME,
    MAESTRA_FILENAME,
    get_catalogs_dir,
    load_config,
    save_config,
)
from app.repositories.homologacion_repo import count_homologacion
from app.repositories.cartera_repo import count_cartera
from app.repositories.maestra_repo import count_maestra
from app.repositories.licitaciones_repo import count_licitaciones
from .schemas import CatalogStatsOut, CatalogImportOut, PrivateHoldingCatalogOut, CarteraSearchOut

router = APIRouter(prefix="/api/catalogs", tags=["catalogs"])


def _save_upload(file: UploadFile, filename: str) -> Path:
    """Guarda el archivo subido en catalogs/ y retorna la ruta."""
    dest = get_catalogs_dir() / filename
    with dest.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    return dest


def _persist_catalog_path(field_name: str, dest: Path) -> None:
    """Persiste en settings.json la ruta final del catálogo subido."""
    cfg = load_config()
    setattr(cfg, field_name, str(dest))
    save_config(cfg)


@router.get("/stats", response_model=CatalogStatsOut)
def catalog_stats():
    try:
        homo = count_homologacion()
        return CatalogStatsOut(
            homologacion_cm=homo["cm"],
            homologacion_sap=homo["sap"],
            maestra=count_maestra(),
            cartera=count_cartera(),
            licitaciones=count_licitaciones(),
            redsalud=_count_redsalud(),
        )
    except Exception as e:
        raise HTTPException(500, detail=str(e))


def _count_redsalud() -> int:
    try:
        from app.services.redsalud_homo_service import get_redsalud_homo_service
        return get_redsalud_homo_service().count()
    except Exception:
        return 0


@router.post("/homologacion", response_model=CatalogImportOut)
async def upload_homologacion(file: UploadFile = File(...)):
    dest = _save_upload(file, HOMOLOGACION_FILENAME)
    from app.services.homologacion_service import get_homologacion_service
    svc = get_homologacion_service()
    cnt, errors = svc.cargar_homologacion_excel(str(dest))
    svc.reload()
    _persist_catalog_path("homologacion_path", dest)
    return CatalogImportOut(imported=cnt, errors=errors)


@router.post("/maestra", response_model=CatalogImportOut)
async def upload_maestra(file: UploadFile = File(...)):
    dest = _save_upload(file, MAESTRA_FILENAME)
    from app.services.homologacion_service import get_homologacion_service
    from app.services.maestra_service import get_maestra_service
    svc_homo = get_homologacion_service()
    cnt, errors = svc_homo.cargar_maestra_sap(str(dest))
    svc_homo.reload()
    svc_maestra = get_maestra_service()
    svc_maestra.cargar_excel(str(dest))
    svc_maestra.reload()
    _persist_catalog_path("maestra_path", dest)
    return CatalogImportOut(imported=cnt, errors=errors)


@router.get("/maestra/search")
def search_maestra_endpoint(q: str):
    if not q or len(q) < 3:
        return []
    from app.repositories.maestra_repo import search_maestra
    return search_maestra(q)


@router.post("/cartera", response_model=CatalogImportOut)
async def upload_cartera(file: UploadFile = File(...)):
    dest = _save_upload(file, CARTERA_FILENAME)
    from app.services.cartera_service import get_cartera_service
    svc = get_cartera_service()
    cnt, errors = svc.cargar_cartera_excel(str(dest))
    svc.reload()
    _persist_catalog_path("cartera_path", dest)
    return CatalogImportOut(imported=cnt, errors=errors)


@router.get("/cartera/search", response_model=list[CarteraSearchOut])
def search_cartera_endpoint(q: str = Query(default="", min_length=2), limit: int = Query(default=8, ge=1, le=20)):
    from app.services.cartera_service import get_cartera_service

    items = get_cartera_service().search(q, limit=limit)
    return [
        CarteraSearchOut(
            cod_cliente=item.cod_cliente,
            rut=item.rut,
            razon=item.razon,
            comuna=item.comuna,
            region_nombre=item.region_nombre,
            cartera=item.cartera,
            vendedor=item.vendedor,
        )
        for item in items
    ]


@router.post("/correos", response_model=CatalogImportOut)
async def upload_correos(file: UploadFile = File(...)):
    dest = _save_upload(file, CORREOS_FILENAME)
    from app.services.email_service import get_email_service
    ok, msg = get_email_service().cargar_correos(str(dest))
    if ok:
        _persist_catalog_path("correos_path", dest)
    return CatalogImportOut(imported=1 if ok else 0, errors=[] if ok else [msg])


@router.post("/redsalud", response_model=CatalogImportOut)
async def upload_redsalud(file: UploadFile = File(...)):
    dest = _save_upload(file, REDSALUD_HOMO_FILENAME)
    from app.services.redsalud_homo_service import get_redsalud_homo_service
    from app.services.private_catalog_service import import_private_catalog

    svc = get_redsalud_homo_service()
    cnt, errors = svc.cargar_excel(str(dest))
    svc.reload()
    import_private_catalog("redsalud", str(dest), file.filename or REDSALUD_HOMO_FILENAME)
    _persist_catalog_path("redsalud_homo_path", dest)
    return CatalogImportOut(imported=cnt, errors=errors)


@router.post("/licitaciones", response_model=CatalogImportOut)
async def upload_licitaciones(file: UploadFile = File(...)):
    dest = _save_upload(file, LICITACIONES_FILENAME)
    from app.services.licitaciones_service import get_licitaciones_service
    svc = get_licitaciones_service()
    cnt, errors = svc.importar_lic(str(dest))
    svc.reload()
    _persist_catalog_path("licitaciones_path", dest)
    return CatalogImportOut(imported=cnt, errors=errors)


@router.get("/private-holdings", response_model=list[PrivateHoldingCatalogOut])
def list_private_holdings_endpoint():
    from app.services.private_catalog_service import list_private_holdings

    return [
        PrivateHoldingCatalogOut(
            id=item.id,
            nombre=item.nombre,
            prefijo=item.prefijo,
            parser_type=item.parser_type,
            homo_file=item.homo_file,
            catalog_count=item.catalog_count,
        )
        for item in list_private_holdings()
    ]


_CATALOG_TEMPLATES: dict[str, dict] = {
    "homologacion": {
        "filename": "plantilla_homologacion.xlsx",
        "columns": ["codigo_cm", "descripcion_cm", "itemcode_sap", "descripcion_sap", "familia", "subfamilia", "unidad_medida"],
        "example": ["CM-001", "GUANTE LATEX TALLA M", "KNE00001", "GUANTE LATEX M NEMO", "Insumos", "Proteccion", "PAR"],
    },
    "maestra": {
        "filename": "plantilla_maestra_sap.xlsx",
        "columns": ["itemcode_sap", "descripcion_sap", "familia", "subfamilia", "unidad_medida", "precio_referencia"],
        "example": ["KNE00001", "GUANTE LATEX M NEMO", "Insumos", "Proteccion", "PAR", "1500"],
    },
    "cartera": {
        "filename": "plantilla_cartera.xlsx",
        "columns": ["cod_cliente", "rut", "razon", "comuna", "region_nombre", "cartera", "vendedor"],
        "example": ["C001", "12.345.678-9", "HOSPITAL EJEMPLO", "Santiago", "RM", "Cartera Norte", "Juan Perez"],
    },
    "correos": {
        "filename": "plantilla_correos.xlsx",
        "columns": ["email", "nombre", "holding", "activo"],
        "example": ["oc@hospital.cl", "Hospital Ejemplo", "redsalud", "si"],
    },
    "redsalud": {
        "filename": "plantilla_homo_redsalud.xlsx",
        "columns": ["codigo_redsalud", "descripcion_redsalud", "itemcode_sap", "descripcion_sap"],
        "example": ["RS-001", "GUANTE LATEX MEDIANO", "KNE00001", "GUANTE LATEX M NEMO"],
    },
    "licitaciones": {
        "filename": "plantilla_licitaciones.xlsx",
        "columns": ["descripcion_comprador", "descripcion_norm", "itemcode_sap", "descripcion_nemo", "rut_comprador", "frecuencia"],
        "example": ["Guante látex mediano", "guante latex mediano", "KNE00001", "GUANTE LATEX M NEMO", "61.002.172-4", "1"],
    },
}


@router.get("/template/{catalog_type}")
def download_catalog_template(catalog_type: str):
    if catalog_type not in _CATALOG_TEMPLATES:
        raise HTTPException(404, detail=f"Tipo de catalogo desconocido: {catalog_type}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, detail="openpyxl no disponible")

    tmpl = _CATALOG_TEMPLATES[catalog_type]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")

    for col_idx, col_name in enumerate(tmpl["columns"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(18, len(col_name) + 4)

    for col_idx, value in enumerate(tmpl["example"], start=1):
        ws.cell(row=2, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{tmpl['filename']}\""},
    )


@router.get("/aprendizaje/export")
def export_aprendizaje():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(500, detail="openpyxl no disponible")

    conn = None
    try:
        from app.db import get_connection
        conn = get_connection()
        rows = conn.execute("""
            SELECT descripcion_comprador, itemcode_sap, descripcion_nemo,
                   rut_comprador, frecuencia
            FROM licitaciones_ref
            ORDER BY frecuencia DESC, descripcion_norm
        """).fetchall()
    finally:
        if conn:
            conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aprendizaje"

    headers = ["descripcion_comprador", "itemcode_sap", "descripcion_nemo", "rut_comprador", "frecuencia"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(20, len(h) + 4)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["descripcion_comprador"] or "")
        ws.cell(row=row_idx, column=2, value=row["itemcode_sap"] or "")
        ws.cell(row=row_idx, column=3, value=row["descripcion_nemo"] or "")
        ws.cell(row=row_idx, column=4, value=row["rut_comprador"] or "")
        ws.cell(row=row_idx, column=5, value=int(row["frecuencia"] or 1))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"aprendizaje_nemooc.xlsx\""},
    )


@router.post("/aprendizaje/import", response_model=CatalogImportOut)
async def import_aprendizaje(file: UploadFile = File(...)):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, detail="openpyxl no disponible")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(c or "").strip().lower() for c in header_row]

        col = {}
        for i, h in enumerate(headers):
            if "descripcion_comprador" in h:
                col["desc"] = i
            elif "itemcode" in h:
                col["itemcode"] = i
            elif "descripcion_nemo" in h:
                col["nemo"] = i
            elif "rut" in h:
                col["rut"] = i
            elif "frecuencia" in h:
                col["freq"] = i

        if "desc" not in col or "itemcode" not in col:
            return CatalogImportOut(imported=0, errors=["Columnas requeridas no encontradas: descripcion_comprador, itemcode_sap"])

        from app.repositories.licitaciones_repo import upsert_from_assignment

        imported = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                desc = str(row[col["desc"]] or "").strip()
                itemcode = str(row[col["itemcode"]] or "").strip()
                if not desc or not itemcode:
                    continue
                nemo = str(row[col["nemo"]] or "").strip() if "nemo" in col else ""
                rut = str(row[col["rut"]] or "").strip() if "rut" in col else ""
                freq = int(row[col["freq"]] or 1) if "freq" in col else 1
                for _ in range(max(1, freq)):
                    upsert_from_assignment(
                        descripcion_comprador=desc,
                        itemcode_sap=itemcode,
                        rut_comprador=rut,
                        descripcion_nemo=nemo,
                    )
                imported += 1
            except Exception as e:
                errors.append(str(e))

        wb.close()
        from app.services.licitaciones_service import get_licitaciones_service
        get_licitaciones_service().reload()
        return CatalogImportOut(imported=imported, errors=errors[:5])

    except Exception as e:
        raise HTTPException(400, detail=f"Error leyendo archivo: {e}")


@router.post("/private/{holding_id}", response_model=CatalogImportOut)
async def upload_private_catalog(holding_id: str, file: UploadFile = File(...)):
    from app.services.private_catalog_service import import_private_catalog, list_private_holdings

    holdings = {item.id: item for item in list_private_holdings()}
    if holding_id not in holdings:
        raise HTTPException(404, detail=f"Holding no existe: {holding_id}")

    filename = holdings[holding_id].homo_file or f"HOMO_{holding_id.upper()}.xlsx"
    dest = _save_upload(file, filename)
    imported, errors = import_private_catalog(holding_id, str(dest), file.filename or filename)
    return CatalogImportOut(imported=imported, errors=errors)
