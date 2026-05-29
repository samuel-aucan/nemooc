"""
Endpoints REST para Órdenes de Compra.
"""
import gzip
import logging
import re
import sys
from pathlib import Path
from typing import Optional, List
from datetime import datetime

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, Response, StreamingResponse

_nemo_oc_dir = Path(__file__).parent.parent.parent.parent / "nemo_oc"
if str(_nemo_oc_dir) not in sys.path:
    sys.path.insert(0, str(_nemo_oc_dir))

from backend.core.repo_selector import oc_repo as oc_repository
from app.config import get_data_dir, load_config
from app.services.sync_service import refresh_oc_status_from_portal
from app.utils.clipboard_utils import generar_texto_sap
from app.services.cartera_service import get_cartera_service
from app.services.licitaciones_service import get_licitaciones_service
from app.services.mp_api_service import APIError, MercadoPublicoAPI
from app.services.mp_portal_service import get_public_oc_portal_meta
from app.services.sap_mode_service import (
    enrich_linea_for_api,
    get_line_sap_values_history,
    reset_line_sap_values,
    update_line_mode,
    update_line_sap_values,
)
from app.services.html_pdf_service import prepare_html_document, render_html_to_pdf
from backend.api.sync_routes import GLOBAL_SYNC_LOGS
from backend.core.auth import get_current_user, get_user_by_id, is_auth_disabled, list_users

from .schemas import (
    OrdenCompraOut, LineaOCOut, OcDetailOut, StatsOut, FiltrosOut, HoldingFiltroOut,
    SapTextOut, AsignarItemcodeIn, SapModeIn, SapValuesIn, SapValuesOut, SapValuesHistoryOut,
    EstadoIn, IngresadaIn, ResponsableIn, ResponsableOut, NotasIn, SugerenciaOut, EstadoHistorialOut,
    CatalogImportOut, AnalyticsOut, AnalyticsSummaryOut, AnalyticsDailyPointOut, AnalyticsRankingItemOut, ReviewQueueItemOut,
    AuditoriaResponse, OcAuditoriaItem, DocumentoFuenteOut, AnalyticsSyncHealthOut,
    ImportarOcMpIn, ImportarOcMpOut,
)

router = APIRouter(prefix="/api/v1/ocs", tags=["ocs"])
logger = logging.getLogger(__name__)


def _enrich_oc(oc, holdings_map: dict | None = None) -> dict:
    """Añade campos de cartera y holding a una OC."""
    d = oc.__dict__.copy()
    cartera_svc = get_cartera_service()
    cliente = cartera_svc.lookup(oc.cliente_sap_sugerido) if oc.cliente_sap_sugerido else None
    d["cartera"]       = cliente.cartera if cliente else ""
    d["vendedor"]      = cliente.vendedor if cliente else ""
    d["region_nombre"] = cliente.region_nombre if cliente else ""
    d["razon_social"]  = cliente.razon if cliente else ""
    if holdings_map is None:
        holdings_map = oc_repository.get_holdings_map()
    holding_id = oc.codigo_organismo or ""
    holding_nombre = holdings_map.get(holding_id, "") if holding_id else ""

    if oc.tipo_oc == "PRIVADA" and not holding_nombre:
        try:
            from app.services.private_holding_service import detect_holding_from_identity

            detection = detect_holding_from_identity(
                rut_value=oc.rut_unidad,
                buyer_name=oc.nombre_organismo,
            )
            if detection.resolved:
                holding_id = holding_id or detection.holding_id
                holding_nombre = detection.holding_nombre
                d["codigo_organismo"] = holding_id
        except Exception:
            pass

    d["holding_nombre"] = holding_nombre
    return d


def _actor_from_request(request: Request) -> tuple[int | None, str]:
    user = get_current_user(request)
    if user.get("auth_disabled"):
        return 0, "Sistema local"
    user_id = int(user.get("id") or 0)
    username = (user.get("username") or user.get("nombre_completo") or "").strip()
    return user_id, username


def _sync_health_payload() -> AnalyticsSyncHealthOut:
    from backend.api.sync_routes import _active
    from backend.core.tasks import (
        get_last_successful_sync_time,
        get_next_light_sync_time,
        get_next_scheduled_sync_time,
    )

    recent_errors = sum(
        1
        for entry in GLOBAL_SYNC_LOGS[-50:]
        if "error" in (entry.get("message") or "").lower()
    )
    last_sync = get_last_successful_sync_time()
    next_sync = get_next_scheduled_sync_time()
    next_light = get_next_light_sync_time()
    return AnalyticsSyncHealthOut(
        running=len(_active) > 0,
        active_tasks=list(_active.keys()),
        last_mp_sync_at=last_sync.isoformat() if last_sync else None,
        next_sync_at=next_sync.isoformat() if next_sync else None,
        next_light_sync=next_light.isoformat() if next_light else None,
        errores_recientes=recent_errors,
    )


def _looks_like_public_oc(codigo_oc: str) -> bool:
    return bool(re.search(r"-[A-Z]{2,4}\d{2}$", (codigo_oc or "").upper()))


def _looks_generic_address(value: str) -> bool:
    normalized = (value or "").strip().lower()
    return normalized in {
        "",
        "sin dirección registrada para unidad de compra",
        "sin direccion registrada para unidad de compra",
        "bienes y servicios",
    }


def _actualizar_campos_publicos_compat(codigo_oc: str, **campos) -> None:
    try:
        oc_repository.actualizar_campos_publicos(codigo_oc, **campos)
    except TypeError:
        oc_repository.actualizar_campos_publicos(codigo_oc, campos)


def _enrich_public_metadata(oc) -> None:
    if not _looks_like_public_oc(oc.codigo_oc):
        return

    updated = False

    if not (oc.codigo_licitacion or "").strip():
        try:
            cfg = load_config()
            if cfg.api_ticket:
                raw = MercadoPublicoAPI(cfg.api_ticket, cfg.codigo_empresa).obtener_detalle_oc(oc.codigo_oc)
                codigo_licitacion = str(raw.get("CodigoLicitacion", "") or "").strip()
                if codigo_licitacion:
                    oc.codigo_licitacion = codigo_licitacion
                    updated = True
        except Exception:
            pass

    needs_despacho = _looks_generic_address(oc.direccion_despacho)
    needs_facturacion = _looks_generic_address(oc.direccion_facturacion)

    if needs_despacho or needs_facturacion:
        portal_meta = get_public_oc_portal_meta(oc.codigo_oc)
        if portal_meta.direccion_despacho and needs_despacho:
            oc.direccion_despacho = portal_meta.direccion_despacho
            updated = True
        if portal_meta.direccion_facturacion and needs_facturacion:
            oc.direccion_facturacion = portal_meta.direccion_facturacion
            updated = True

    if updated:
        _actualizar_campos_publicos_compat(
            oc.codigo_oc,
            codigo_licitacion=oc.codigo_licitacion or "",
            direccion_despacho=oc.direccion_despacho or "",
            direccion_facturacion=oc.direccion_facturacion or "",
        )


def _resolve_document_snapshot_path(documento: dict) -> Path:
    snapshot_path = (documento.get("snapshot_path") or "").strip()
    if not snapshot_path:
        raise HTTPException(404, detail="La OC no tiene snapshot documental disponible")

    data_dir = get_data_dir().resolve()
    resolved = (data_dir / snapshot_path).resolve()
    if data_dir not in resolved.parents and resolved != data_dir:
        raise HTTPException(400, detail="Ruta documental invalida")
    if not resolved.exists() or not resolved.is_file():
        raise HTTPException(404, detail="No se encontro el respaldo documental")
    return resolved


def _load_document_html(documento: dict) -> str:
    snapshot_type = (documento.get("snapshot_type") or "").strip().lower()
    snapshot_file = _resolve_document_snapshot_path(documento)

    if snapshot_type == "html_gzip":
        with gzip.open(snapshot_file, "rt", encoding="utf-8") as fh:
            return fh.read()
    return snapshot_file.read_text(encoding="utf-8")


def _safe_download_filename(value: str, fallback: str = "documento.pdf") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (value or "").strip())
    return cleaned.strip("._-") or fallback


def _load_imap_pdf_document(documento: dict) -> tuple[bytes, str]:
    from app.services.imap_service import recuperar_pdf_adjunto_imap

    cfg = load_config()
    payload = documento.get("access_payload") or {}
    if not isinstance(payload, dict):
        payload = {}

    try:
        pdf_bytes, filename = recuperar_pdf_adjunto_imap(
            smtp_user=cfg.smtp_user,
            smtp_password=cfg.smtp_password,
            imap_server=cfg.imap_server,
            imap_port=cfg.imap_port,
            imap_folder=payload.get("imap_folder") or cfg.imap_folder,
            imap_uid=str(payload.get("imap_uid") or ""),
            message_id=str(payload.get("message_id") or ""),
            attachment_index=int(payload.get("attachment_index") or 1),
            attachment_filename=str(payload.get("attachment_filename") or ""),
            attachment_sha256=str(payload.get("attachment_sha256") or documento.get("snapshot_sha256") or ""),
        )
    except FileNotFoundError as exc:
        raise HTTPException(404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(400, detail=str(exc))
    except ConnectionError as exc:
        raise HTTPException(502, detail=str(exc))
    except Exception as exc:
        logger.exception("No se pudo recuperar PDF privado por IMAP")
        raise HTTPException(500, detail=f"No se pudo recuperar el PDF desde el correo original: {exc}")

    return pdf_bytes, _safe_download_filename(filename, "orden_compra.pdf")


# ── Lista ────────────────────────────────────────────────────────────────────

def _refresh_single_oc_portal_status(codigo_oc: str) -> bool:
    if not _looks_like_public_oc(codigo_oc):
        return False

    cfg = load_config()
    if not cfg.api_ticket:
        return False

    try:
        result = refresh_oc_status_from_portal(
            ticket=cfg.api_ticket,
            codigo_empresa=cfg.codigo_empresa,
            codigo_oc=codigo_oc,
        )
        return bool(result.get("updated"))
    except Exception as exc:
        logger.warning("No se pudo refrescar estado_mp de %s tras cambio local: %s", codigo_oc, exc)
        return False


@router.get("", response_model=list[OrdenCompraOut])
def list_ocs(
    estado:     List[str] = Query(default=[]),
    estado_mp:  List[str] = Query(default=[]),
    tipo_oc:    List[str] = Query(default=[]),
    cartera:    List[str] = Query(default=[]),
    holding:    List[str] = Query(default=[]),
    responsable: List[str] = Query(default=[]),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    fecha_ingreso_desde: Optional[str] = Query(None),
    fecha_ingreso_hasta: Optional[str] = Query(None),
    busqueda:   Optional[str] = Query(None),
):
    ocs = oc_repository.get_all_ocs(
        estado=estado or None,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_ingreso_desde=fecha_ingreso_desde,
        fecha_ingreso_hasta=fecha_ingreso_hasta,
        busqueda=busqueda,
        estado_mp=estado_mp or None,
        tipo_oc=tipo_oc or None,
        holding=holding or None,
        responsable=responsable or None,
    )
    holdings_map = oc_repository.get_holdings_map()
    enriched = [OrdenCompraOut(**_enrich_oc(oc, holdings_map)) for oc in ocs]
    if cartera:
        enriched = [o for o in enriched if o.cartera in cartera]
    return enriched


@router.get("/stats", response_model=StatsOut)
def get_stats():
    """Conteo rápido: total, sin homologar, ingresadas."""
    return StatsOut(**oc_repository.get_stats())


@router.get("/filtros", response_model=FiltrosOut)
def get_filtros():
    cartera_svc = get_cartera_service()
    carteras = cartera_svc.list_carteras()

    raw_holdings = oc_repository.get_distinct_holdings()
    holdings = [HoldingFiltroOut(id=h["id"], nombre=h["nombre"]) for h in raw_holdings]

    return FiltrosOut(
        estados_mp=oc_repository.get_distinct_estados_mp(),
        tipos=oc_repository.get_distinct_tipos(),
        carteras=carteras,
        holdings=holdings,
    )


@router.get("/analytics", response_model=AnalyticsOut)
def get_analytics(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    limit: int = Query(150, ge=20, le=400),
):
    cartera_svc = get_cartera_service()
    licitaciones_svc = get_licitaciones_service()

    summary = oc_repository.get_analytics_summary(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    received_by_day = oc_repository.get_received_by_day(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    entered_by_day = oc_repository.get_entered_by_day(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    top_clients = oc_repository.get_top_clients_by_amount(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    top_buyers = oc_repository.get_top_buyers_by_amount(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    control = oc_repository.get_control_analytics(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    total_cola_sin_limite = oc_repository.get_review_queue_count(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    queue_rows = oc_repository.get_review_queue(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        limit=limit,
    )

    # Batch lookup: evitar N+1 queries (antes: 150 lookups, ahora: 1)
    cliente_codes = [r["cliente_sap_sugerido"] for r in queue_rows if r["cliente_sap_sugerido"]]
    clientes_map = cartera_svc.lookup_batch(cliente_codes) if cliente_codes else {}

    queue: list[ReviewQueueItemOut] = []

    for row in queue_rows:
        cliente = clientes_map.get(row["cliente_sap_sugerido"]) if row["cliente_sap_sugerido"] else None
        is_pending = (row["estado_homologacion"] or "pendiente") == "pendiente" or not (row["itemcode_sap"] or "").strip()

        sugerencia_principal = None
        if is_pending:
            texto = " ".join(filter(None, [row["especificacion_comprador"], row["producto"]])).strip()
            if texto:
                sugerencias = licitaciones_svc.buscar_sugerencias(
                    texto,
                    rut_oc=row["rut_unidad"],
                    max_results=1,
                )
                if sugerencias:
                    top = sugerencias[0]
                    sugerencia_principal = SugerenciaOut(
                        itemcode_sap=top.itemcode_sap,
                        descripcion_sap=top.descripcion_sap,
                        descripcion_match=top.descripcion_match,
                        frecuencia=top.frecuencia,
                        score=top.score,
                        estrellas=max(1, round(top.score * 5)),
                    )

        queue.append(
            ReviewQueueItemOut(
                codigo_oc=row["codigo_oc"],
                correlativo=row["correlativo"],
                fecha_envio=row["fecha_envio"] or "",
                tipo_oc=row["tipo_oc"] or "",
                nombre_organismo=row["nombre_organismo"] or "",
                cliente_sap_sugerido=row["cliente_sap_sugerido"] or "",
                cartera=cliente.cartera if cliente else "",
                estado_interno=row["estado_interno"] or "Nueva",
                estado_homologacion=row["estado_homologacion"] or "pendiente",
                itemcode_sap=row["itemcode_sap"],
                descripcion_sap=row["descripcion_sap"],
                producto=row["producto"] or "",
                especificacion_comprador=row["especificacion_comprador"] or "",
                cantidad=float(row["cantidad"] or 0),
                total=float(row["total"] or 0),
                rut_unidad=row["rut_unidad"] or "",
                sugerencia_principal=sugerencia_principal,
            )
        )

    return AnalyticsOut(
        summary=AnalyticsSummaryOut(
            **summary,
            cola_revision=len(queue),
            total_cola_sin_limite=total_cola_sin_limite,
        ),
        received_by_day=[AnalyticsDailyPointOut(**row) for row in received_by_day],
        entered_by_day=[AnalyticsDailyPointOut(**row) for row in entered_by_day],
        top_clients=[AnalyticsRankingItemOut(**row) for row in top_clients],
        top_buyers=[AnalyticsRankingItemOut(**row) for row in top_buyers],
        queue=queue,
        productividad_hoy=control["productividad_hoy"],
        productividad_usuarios=control["productividad_usuarios"],
        aging=control["aging"],
        funnel=control["funnel"],
        salud_sync=_sync_health_payload(),
        privadas=control["privadas"],
        top_blockers=control["top_blockers"],
    )


@router.get("/auditoria", response_model=AuditoriaResponse)
def get_auditoria(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
):
    # Auditoria queda fija desde hoy hacia adelante.
    fecha_desde = datetime.now().date().isoformat()
    fecha_hasta = None
    data = oc_repository.get_auditoria(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    return AuditoriaResponse(
        aceptadas_sin_ingresar=[OcAuditoriaItem(**r) for r in data["aceptadas_sin_ingresar"]],
        ingresadas_sin_aceptar=[OcAuditoriaItem(**r) for r in data["ingresadas_sin_aceptar"]],
    )


@router.get("/export-all")
def export_all(
    estado:     List[str] = Query(default=[]),
    estado_mp:  List[str] = Query(default=[]),
    tipo_oc:    List[str] = Query(default=[]),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    busqueda:   Optional[str] = Query(None),
):
    import openpyxl
    from io import BytesIO

    ocs = oc_repository.get_all_ocs(
        estado=estado or None, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        busqueda=busqueda, estado_mp=estado_mp or None, tipo_oc=tipo_oc or None,
    )
    cartera_svc = get_cartera_service()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCs"

    headers = [
        "Código OC", "Tipo", "Estado MP", "Estado Interno",
        "Fecha Envío", "Comprador", "Cliente SAP", "Cartera",
        "Ejecutivo", "Total Neto", "Impuestos", "Total", "Moneda",
        "Líneas", "Notas"
    ]
    ws.append(headers)

    for oc in ocs:
        cliente = cartera_svc.lookup(oc.cliente_sap_sugerido) if oc.cliente_sap_sugerido else None
        ws.append([
            oc.codigo_oc, oc.tipo_oc, oc.estado_mp, oc.estado_interno,
            oc.fecha_envio[:10] if oc.fecha_envio else "",
            oc.nombre_organismo, oc.cliente_sap_sugerido,
            cliente.cartera if cliente else "",
            cliente.vendedor if cliente else "",
            oc.total_neto, oc.impuestos, oc.total, oc.moneda,
            oc.cantidad_lineas, oc.notas or "",
        ])

    for row in ws.iter_rows(min_row=2, min_col=10, max_col=12):
        for cell in row:
            cell.number_format = '#,##0.####'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=OCs_NemoChile_{fecha}.xlsx"},
    )


# ── Detalle ──────────────────────────────────────────────────────────────────

@router.get("/responsables", response_model=list[ResponsableOut])
def get_responsables():
    if is_auth_disabled():
        return [ResponsableOut(id=0, username="Sistema local", nombre_completo="Sistema local")]
    return [
        ResponsableOut(
            id=int(user["id"]),
            username=user["username"],
            nombre_completo=user.get("nombre_completo") or "",
        )
        for user in list_users()
        if user.get("activo", True)
    ]


@router.get("/{codigo_oc}", response_model=OcDetailOut)
def get_oc(codigo_oc: str):
    oc = oc_repository.get_oc(codigo_oc)
    if not oc:
        raise HTTPException(404, detail=f"OC {codigo_oc} no encontrada")
    try:
        _enrich_public_metadata(oc)
    except Exception:
        pass  # non-critical enrichment
    lineas_raw = oc_repository.get_lineas(codigo_oc)
    lineas = []
    for linea in lineas_raw:
        try:
            lineas.append(enrich_linea_for_api(linea, oc.tipo_oc))
        except Exception:
            lineas.append(linea)
    historial_estados = oc_repository.get_estado_historial(codigo_oc)
    documento = oc_repository.get_document_source(codigo_oc)
    return OcDetailOut(
        cabecera=OrdenCompraOut(**_enrich_oc(oc)),
        lineas=[LineaOCOut(**l.__dict__) for l in lineas],
        historial_estados=[EstadoHistorialOut(**h) for h in historial_estados],
        documento=DocumentoFuenteOut(**documento) if documento else None,
    )


@router.get("/{codigo_oc}/documento")
def get_oc_documento(
    codigo_oc: str,
    download: bool = Query(False),
    auto_print: bool = Query(False),
):
    documento = oc_repository.get_document_source(codigo_oc)
    if not documento or not documento.get("document_available"):
        raise HTTPException(404, detail=f"La OC {codigo_oc} no tiene documento disponible")

    if (documento.get("source_type") or "").strip().lower() == "imap_attachment":
        pdf_bytes, filename = _load_imap_pdf_document(documento)
        disposition = "attachment" if download else "inline"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
        )

    html = prepare_html_document(
        _load_document_html(documento),
        source_locator=documento.get("source_locator", ""),
        auto_print=auto_print,
    )

    headers = {}
    if download:
        headers["Content-Disposition"] = f'attachment; filename="{codigo_oc}.html"'

    return HTMLResponse(content=html, headers=headers)


@router.get("/{codigo_oc}/documento-pdf")
def get_oc_documento_pdf(codigo_oc: str):
    documento = oc_repository.get_document_source(codigo_oc)
    if not documento or not documento.get("document_available"):
        raise HTTPException(404, detail=f"La OC {codigo_oc} no tiene documento disponible")

    if (documento.get("source_type") or "").strip().lower() == "imap_attachment":
        pdf_bytes, filename = _load_imap_pdf_document(documento)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    try:
        pdf_bytes = render_html_to_pdf(
            _load_document_html(documento),
            source_locator=documento.get("source_locator", ""),
            filename_hint=codigo_oc,
        )
    except RuntimeError as exc:
        raise HTTPException(503, detail=str(exc))
    except Exception as exc:
        logger.exception("No se pudo generar PDF para %s", codigo_oc)
        raise HTTPException(500, detail=f"No se pudo generar el PDF automaticamente: {exc}")

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{codigo_oc}.pdf"'},
    )


@router.put("/{codigo_oc}/estado")
def update_estado(codigo_oc: str, body: EstadoIn, request: Request):
    actor_user_id, actor_username = _actor_from_request(request)
    oc_repository.actualizar_estado(
        codigo_oc,
        body.estado,
        actor_user_id=actor_user_id,
        actor_username=actor_username,
    )
    portal_status_refreshed = _refresh_single_oc_portal_status(codigo_oc)
    # Dual-write: sincronizar estado_interno en Supabase
    try:
        from backend.supabase_write_service import sync_estado_oc as _sync_sb
        _sync_sb(codigo_oc, {"estado_interno": body.estado})
    except Exception:
        pass
    return {"ok": True, "portal_status_refreshed": portal_status_refreshed}


@router.put("/{codigo_oc}/ingresada")
def marcar_ingresada(codigo_oc: str, request: Request, body: IngresadaIn | None = None):
    actor_user_id, actor_username = _actor_from_request(request)
    acuerdo = bool(body.acuerdo_global) if body else False
    oc_repository.marcar_ingresada(
        codigo_oc,
        actor_user_id=actor_user_id,
        actor_username=actor_username,
        acuerdo_global=acuerdo,
    )
    portal_status_refreshed = _refresh_single_oc_portal_status(codigo_oc)
    # Dual-write: sincronizar estado_interno + acuerdo_global en Supabase
    try:
        from backend.supabase_write_service import sync_estado_oc as _sync_sb
        _sync_sb(codigo_oc, {
            "estado_interno": "Ingresada",
            "ingreso_sap_acuerdo_global": acuerdo,
            "ingresado_por_username": actor_username or None,
        })
    except Exception:
        pass
    return {"ok": True, "portal_status_refreshed": portal_status_refreshed}


@router.put("/{codigo_oc}/responsable")
def update_responsable(codigo_oc: str, body: ResponsableIn):
    username = ""
    user_id = body.user_id
    if user_id is not None:
        if user_id == 0 and is_auth_disabled():
            username = "Sistema local"
        else:
            user = get_user_by_id(user_id)
            if not user or not user.get("activo", 1):
                raise HTTPException(404, detail="Usuario responsable no encontrado")
            username = user.get("username") or ""

    oc_repository.asignar_responsable_ingreso(codigo_oc, user_id, username)
    # Dual-write: sincronizar responsable_ingreso_username en Supabase
    try:
        from backend.supabase_write_service import sync_estado_oc as _sync_sb
        _sync_sb(codigo_oc, {"responsable_ingreso_username": username or None})
    except Exception:
        pass
    return {"ok": True}


@router.put("/{codigo_oc}/notas")
def update_notas(codigo_oc: str, body: NotasIn):
    oc_repository.guardar_notas(codigo_oc, body.notas)
    # Dual-write: sincronizar notas en Supabase
    try:
        from backend.supabase_write_service import sync_estado_oc as _sync_sb
        _sync_sb(codigo_oc, {"notas": body.notas})
    except Exception:
        pass
    return {"ok": True}


@router.post("/{codigo_oc}/refresh-mp-status")
def refresh_mp_status(codigo_oc: str):
    codigo = (codigo_oc or "").strip().upper()
    if not codigo:
        raise HTTPException(400, detail="Debe indicar un codigo de OC")
    if not _looks_like_public_oc(codigo):
        raise HTTPException(400, detail="La actualizacion MP solo aplica a OCs de Mercado Publico")

    cfg = load_config()
    if not cfg.api_ticket:
        raise HTTPException(400, detail="API ticket no configurado")

    try:
        result = refresh_oc_status_from_portal(
            ticket=cfg.api_ticket,
            codigo_empresa=cfg.codigo_empresa,
            codigo_oc=codigo,
        )
    except APIError as exc:
        status_code = 404 if "sin detalle" in str(exc).lower() else 502
        raise HTTPException(status_code, detail=str(exc))
    except Exception as exc:
        logger.exception("No se pudo refrescar estado MP de %s", codigo)
        raise HTTPException(502, detail=f"No se pudo consultar Mercado Publico: {exc}")

    try:
        from backend.core.tasks import record_mp_sync_success

        record_mp_sync_success()
    except Exception:
        pass

    return {
        "ok": True,
        "updated": bool(result.get("updated")),
        "estado_mp": str(result.get("estado_mp") or ""),
        "codigo_estado_mp": int(result.get("codigo_estado_mp") or 0),
        "refreshed_at": datetime.now().isoformat(),
    }


@router.get("/{codigo_oc}/sap-text", response_model=SapTextOut)
def get_sap_text(codigo_oc: str):
    lineas = oc_repository.get_lineas(codigo_oc)
    texto, excluidos = generar_texto_sap(lineas)
    return SapTextOut(text=texto, excluidos=excluidos)


@router.post("/importar-mp", response_model=ImportarOcMpOut)
def importar_oc_mp(body: ImportarOcMpIn):
    codigo_oc = (body.codigo_oc or "").strip().upper()
    if not codigo_oc:
        raise HTTPException(400, detail="Debe indicar un codigo de OC")

    cfg = load_config()
    if not cfg.api_ticket:
        raise HTTPException(400, detail="API ticket no configurado")

    try:
        from app.services.sync_service import import_single_public_oc

        result = import_single_public_oc(
            ticket=cfg.api_ticket,
            codigo_empresa=cfg.codigo_empresa,
            codigo_oc=codigo_oc,
        )
    except APIError as exc:
        status_code = 404 if "sin detalle" in str(exc).lower() else 502
        raise HTTPException(status_code, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(400, detail=str(exc))
    except Exception as exc:
        logger.exception("No se pudo importar OC puntual desde MP")
        raise HTTPException(500, detail=f"No se pudo importar la OC desde Mercado Publico: {exc}")

    oc = result["oc"]
    return ImportarOcMpOut(
        ok=True,
        created=bool(result.get("created")),
        codigo_oc=oc.codigo_oc,
        message=str(result.get("message") or ""),
        oc=OrdenCompraOut(**_enrich_oc(oc)),
    )


@router.get("/{codigo_oc}/export-excel")
def export_excel(codigo_oc: str):
    import openpyxl
    from io import BytesIO

    oc = oc_repository.get_oc(codigo_oc)
    if not oc:
        raise HTTPException(404)
    lineas = oc_repository.get_lineas(codigo_oc)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = codigo_oc

    ws.append(["Código OC", oc.codigo_oc])
    ws.append(["Tipo", oc.tipo_oc])
    ws.append(["Comprador", oc.nombre_organismo])
    ws.append(["Cliente SAP", oc.cliente_sap_sugerido])
    ws.append(["Total", oc.total, oc.moneda])
    ws.append([])

    headers = [
        "#", "Producto", "Espec. Comprador", "Cód MP",
        "ItemCode SAP", "Desc SAP", "Cantidad", "Cant SAP",
        "Precio Neto", "Precio SAP", "Total", "Estado"
    ]
    ws.append(headers)
    for l in lineas:
        ws.append([
            l.correlativo, l.producto, l.especificacion_comprador, l.codigo_mp or "",
            l.itemcode_sap or "", l.descripcion_sap or "",
            l.cantidad, l.cantidad_sap,
            l.precio_neto, l.precio_sap,
            l.total, l.estado_homologacion,
        ])

    for row in ws.iter_rows(min_row=8, min_col=7, max_col=11):
        for cell in row:
            cell.number_format = '#,##0.####'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={codigo_oc}.xlsx"},
    )


# ── Líneas ───────────────────────────────────────────────────────────────────

@router.put("/{codigo_oc}/lineas/{correlativo}/asignar")
def asignar_itemcode(codigo_oc: str, correlativo: int, body: AsignarItemcodeIn):
    estado = 'sugerido' if body.origen == 'sugerencia' else 'manual'
    oc_repository.asignar_itemcode_linea(
        codigo_oc=codigo_oc,
        correlativo=correlativo,
        itemcode_sap=body.itemcode_sap,
        descripcion_sap=body.descripcion_sap,
        origen=estado,
    )
    # Retornar los datos actualizados para que el frontend los muestre
    return {
        "ok": True,
        "itemcode_sap": body.itemcode_sap,
        "descripcion_sap": body.descripcion_sap or "",
        "estado_homologacion": estado,
    }


@router.delete("/{codigo_oc}/lineas/{correlativo}/asignar")
def limpiar_asignacion(codigo_oc: str, correlativo: int):
    oc_repository.limpiar_asignacion_linea(codigo_oc, correlativo)
    return {
        "ok": True,
        "itemcode_sap": None,
        "descripcion_sap": None,
        "estado_homologacion": "pendiente",
    }


@router.put("/{codigo_oc}/lineas/{correlativo}/sap-mode")
def update_line_sap_mode(codigo_oc: str, correlativo: int, body: SapModeIn):
    try:
        update_line_mode(codigo_oc, correlativo, body.mode)
    except ValueError as exc:
        raise HTTPException(400, detail=str(exc))
    return {"ok": True}


@router.put("/{codigo_oc}/lineas/{correlativo}/sap-values", response_model=SapValuesOut)
def update_line_sap_values_endpoint(codigo_oc: str, correlativo: int, body: SapValuesIn, request: Request):
    actor_user_id, actor_username = _actor_from_request(request)
    try:
        result = update_line_sap_values(
            codigo_oc,
            correlativo,
            body.cantidad_sap,
            body.precio_sap,
            actor_user_id=actor_user_id,
            actor_username=actor_username,
        )
    except ValueError as exc:
        raise HTTPException(400, detail=str(exc))

    try:
        from backend.supabase_write_service import sync_homologacion as _sync_sb

        linea = next((l for l in oc_repository.get_lineas(codigo_oc) if l.correlativo == correlativo), None)
        _sync_sb(
            codigo_oc=codigo_oc,
            nro_linea=correlativo,
            itemcode_sap=linea.itemcode_sap if linea else None,
            descripcion_sap=linea.descripcion_sap if linea else None,
            cantidad_sap=result["cantidad_sap"],
            precio_sap=result["precio_sap"],
            sap_mode=linea.sap_mode if linea else None,
            estado_homologacion=linea.estado_homologacion if linea else "pendiente",
        )
    except Exception:
        pass

    return SapValuesOut(**result)


@router.delete("/{codigo_oc}/lineas/{correlativo}/sap-values", response_model=SapValuesOut)
def reset_line_sap_values_endpoint(codigo_oc: str, correlativo: int, request: Request):
    actor_user_id, actor_username = _actor_from_request(request)
    try:
        result = reset_line_sap_values(
            codigo_oc,
            correlativo,
            actor_user_id=actor_user_id,
            actor_username=actor_username,
        )
    except ValueError as exc:
        raise HTTPException(400, detail=str(exc))

    try:
        from backend.supabase_write_service import sync_homologacion as _sync_sb

        linea = next((l for l in oc_repository.get_lineas(codigo_oc) if l.correlativo == correlativo), None)
        _sync_sb(
            codigo_oc=codigo_oc,
            nro_linea=correlativo,
            itemcode_sap=linea.itemcode_sap if linea else None,
            descripcion_sap=linea.descripcion_sap if linea else None,
            cantidad_sap=result["cantidad_sap"],
            precio_sap=result["precio_sap"],
            sap_mode=linea.sap_mode if linea else None,
            estado_homologacion=linea.estado_homologacion if linea else "pendiente",
        )
    except Exception:
        pass

    return SapValuesOut(**result)


@router.get("/{codigo_oc}/lineas/{correlativo}/sap-values/historial", response_model=list[SapValuesHistoryOut])
def get_line_sap_values_history_endpoint(
    codigo_oc: str,
    correlativo: int,
    limit: int = Query(20, ge=1, le=100),
):
    return [SapValuesHistoryOut(**row) for row in get_line_sap_values_history(codigo_oc, correlativo, limit=limit)]


@router.post("/{codigo_oc}/rehomologar-privada")
def rehomologar_privada(codigo_oc: str):
    """Re-ejecuta el lookup del catálogo privado sobre las líneas sin itemcode."""
    from app.services.private_holding_service import lookup_private_catalog

    oc = oc_repository.get_oc(codigo_oc)
    if not oc:
        raise HTTPException(404, "OC no encontrada")
    if oc.tipo_oc != "PRIVADA":
        raise HTTPException(400, "Solo aplica a OCs privadas")

    holding_id = oc.codigo_organismo
    lineas = oc_repository.get_lineas(codigo_oc)
    actualizadas = 0

    for linea in lineas:
        if linea.itemcode_sap or not linea.codigo_mp:
            continue
        homo = lookup_private_catalog(holding_id, linea.codigo_mp)
        if homo and homo.itemcode_sap:
            oc_repository.asignar_itemcode_linea(
                codigo_oc,
                linea.correlativo,
                homo.itemcode_sap,
                homo.descripcion or "",
                origen="homologado",
            )
            actualizadas += 1

    return {"actualizadas": actualizadas}


@router.get("/{codigo_oc}/lineas/{correlativo}/sugerencias", response_model=list[SugerenciaOut])
def get_sugerencias(codigo_oc: str, correlativo: int):
    oc = oc_repository.get_oc(codigo_oc)
    if not oc:
        raise HTTPException(404)
        
    lineas = oc_repository.get_lineas(codigo_oc)
    linea = next((l for l in lineas if l.correlativo == correlativo), None)
    if not linea:
        raise HTTPException(404)

    texto = " ".join(filter(None, [linea.especificacion_comprador, linea.producto]))
    if not texto.strip():
        return []

    svc = get_licitaciones_service()
    sugs = svc.buscar_sugerencias(texto, rut_oc=oc.rut_unidad, max_results=5)
    return [
        SugerenciaOut(
            itemcode_sap=s.itemcode_sap,
            descripcion_sap=s.descripcion_sap,
            descripcion_match=s.descripcion_match,
            frecuencia=s.frecuencia,
            score=s.score,
            estrellas=max(1, round(s.score * 5)),
        )
        for s in sugs
    ]
