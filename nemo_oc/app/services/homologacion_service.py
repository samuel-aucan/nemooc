"""
Servicio de homologación de productos.
Carga HOMOLOGACION.xlsx y el archivo opcional de Maestra SAP.
Construye el catálogo en memoria y lo persiste en SQLite.
"""
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from app.models.homologacion import HomologacionItem, SapArticulo
from app.db import get_connection

logger = logging.getLogger(__name__)

# Columnas de HOMOLOGACION.xlsx (índice base 0)
COL_ID = 0        # A: ID (codigo_mp)
COL_TIPO = 1      # B: TIPO
COL_MARCA = 2     # C: MARCA
COL_NEMO = 3      # D: COD NEMO
COL_SAP = 4       # E: COD SAP (itemcode_sap)
COL_MODELO = 5    # F: MODELO (descripcion_sap)
COL_PRECIO_ID = 6 # G: PRECIO ID
COL_FEMP = 7      # H: F. EMP (factor_empaque)
COL_PRECIO_U = 8  # I: PRECIO UNI
COL_ACTIVO = 9    # J: Activo


class HomologacionService:
    """Gestiona la carga y búsqueda del catálogo de homologación."""

    def __init__(self):
        self._catalog: Dict[str, HomologacionItem] = {}   # codigo_mp → item
        self._sap_desc: Dict[str, str] = {}               # itemcode_sap → descripcion
        self._loaded = False

    # ------------------------------------------------------------------
    # Carga desde Excel
    # ------------------------------------------------------------------

    def cargar_homologacion_excel(self, path: str) -> Tuple[int, List[str]]:
        """
        Lee HOMOLOGACION.xlsx y guarda en SQLite + memoria.
        Retorna (registros_importados, errores).
        """
        try:
            import openpyxl
        except ImportError:
            return 0, ["openpyxl no está instalado."]

        errores: List[str] = []
        items: List[HomologacionItem] = []
        path_obj = Path(path)

        if not path_obj.exists():
            return 0, [f"Archivo no encontrado: {path}"]

        try:
            wb = openpyxl.load_workbook(str(path_obj), data_only=True, read_only=True)
            ws = wb.active
            now = datetime.now().isoformat()
            filename = path_obj.name

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 6:
                    continue
                raw_id = row[COL_ID]
                raw_sap = row[COL_SAP]
                raw_modelo = row[COL_MODELO]
                raw_femp = row[COL_FEMP]

                if raw_id is None or raw_sap is None:
                    continue

                codigo_mp = str(raw_id).strip()
                itemcode_sap = str(raw_sap).strip()
                descripcion = str(raw_modelo).strip() if raw_modelo else ""

                # Factor de empaque: default 1 si no es numérico
                try:
                    femp = float(raw_femp) if raw_femp is not None else 1.0
                    if femp <= 0:
                        femp = 1.0
                except (ValueError, TypeError):
                    femp = 1.0

                if not codigo_mp or not itemcode_sap:
                    continue

                items.append(HomologacionItem(
                    codigo_mp=codigo_mp,
                    itemcode_sap=itemcode_sap,
                    descripcion_sap=descripcion,
                    factor_empaque=femp,
                    activo=True,
                    origen_archivo=filename,
                    created_at=now,
                    updated_at=now,
                ))

            wb.close()
        except Exception as e:
            return 0, [f"Error leyendo Excel: {e}"]

        if not items:
            return 0, ["No se encontraron registros válidos en el archivo."]

        # Persistir en SQLite
        count = self._guardar_homologacion(items)
        # Recargar en memoria
        self._reload_catalog()
        logger.info(f"Homologación cargada: {count} registros desde {filename}")
        return count, errores

    def cargar_maestra_sap(self, path: str) -> Tuple[int, List[str]]:
        """
        Lee el archivo de Maestra SAP (xlsx o csv) y actualiza descripciones.
        El archivo debe tener al menos columnas con ItemCode y Descripción.
        Retorna (registros_importados, errores).
        """
        errores: List[str] = []
        path_obj = Path(path)
        ext = path_obj.suffix.lower()

        if not path_obj.exists():
            return 0, [f"Archivo no encontrado: {path}"]

        rows: List[Tuple[str, str]] = []  # (itemcode, descripcion)

        if ext in ('.xlsx', '.xlsm', '.xls'):
            rows, errores = self._leer_maestra_excel(path_obj)
        elif ext == '.csv':
            rows, errores = self._leer_maestra_csv(path_obj)
        else:
            return 0, [f"Formato no soportado: {ext}. Use .xlsx o .csv"]

        if not rows:
            return 0, errores + ["No se encontraron registros válidos."]

        now = datetime.now().isoformat()
        articulos = [
            SapArticulo(
                itemcode_sap=code,
                descripcion_sap=desc,
                activo=True,
                origen_archivo=path_obj.name,
                created_at=now,
                updated_at=now,
            )
            for code, desc in rows if code
        ]

        count = self._guardar_sap_articulos(articulos)
        self._reload_sap_desc()
        logger.info(f"Maestra SAP cargada: {count} registros desde {path_obj.name}")
        return count, errores

    # ------------------------------------------------------------------
    # Lookup
    # ------------------------------------------------------------------

    def lookup(self, codigo_mp: str) -> Optional[HomologacionItem]:
        """Busca un ítem por codigo_mp. Retorna None si no existe."""
        if not self._loaded:
            self._reload_catalog()
        item = self._catalog.get(str(codigo_mp))
        if item and item.descripcion_sap is None:
            # Intentar enriquecer con descripción SAP si se importó Maestra
            item.descripcion_sap = self._sap_desc.get(item.itemcode_sap)
        return item

    def get_stats(self) -> dict:
        """Retorna estadísticas del catálogo cargado en memoria."""
        if not self._loaded:
            self._reload_catalog()
        return {
            "cm_registros": len(self._catalog),
            "sap_articulos": len(self._sap_desc),
            "cruzados": sum(
                1 for item in self._catalog.values()
                if item.itemcode_sap in self._sap_desc
            ),
        }

    def reload(self):
        """Fuerza recarga desde la base de datos."""
        self._reload_catalog()
        self._reload_sap_desc()

    # ------------------------------------------------------------------
    # Persistencia
    # ------------------------------------------------------------------

    def _guardar_homologacion(self, items: List[HomologacionItem]) -> int:
        conn = get_connection()
        count = 0
        try:
            now = datetime.now().isoformat()
            for item in items:
                conn.execute("""
                    INSERT INTO homologacion_productos
                        (codigo_mp, itemcode_sap, descripcion_sap, factor_empaque,
                         activo, origen_archivo, created_at, updated_at)
                    VALUES (?, ?, ?, ?, 1, ?, ?, ?)
                    ON CONFLICT(codigo_mp) DO UPDATE SET
                        itemcode_sap    = excluded.itemcode_sap,
                        descripcion_sap = excluded.descripcion_sap,
                        factor_empaque  = excluded.factor_empaque,
                        origen_archivo  = excluded.origen_archivo,
                        updated_at      = excluded.updated_at
                """, (
                    item.codigo_mp, item.itemcode_sap, item.descripcion_sap,
                    item.factor_empaque, item.origen_archivo,
                    item.created_at, now
                ))
                count += 1
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"Error guardando homologación: {e}")
            raise
        finally:
            conn.close()
        return count

    def _guardar_sap_articulos(self, articulos: List[SapArticulo]) -> int:
        conn = get_connection()
        count = 0
        try:
            now = datetime.now().isoformat()
            for art in articulos:
                conn.execute("""
                    INSERT INTO sap_articulos
                        (itemcode_sap, descripcion_sap, activo, origen_archivo,
                         created_at, updated_at)
                    VALUES (?, ?, 1, ?, ?, ?)
                    ON CONFLICT(itemcode_sap) DO UPDATE SET
                        descripcion_sap = excluded.descripcion_sap,
                        origen_archivo  = excluded.origen_archivo,
                        updated_at      = excluded.updated_at
                """, (
                    art.itemcode_sap, art.descripcion_sap,
                    art.origen_archivo, art.created_at, now
                ))
                count += 1
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"Error guardando artículos SAP: {e}")
            raise
        finally:
            conn.close()
        return count

    def _reload_catalog(self):
        conn = get_connection()
        try:
            rows = conn.execute("""
                SELECT codigo_mp, itemcode_sap, descripcion_sap, factor_empaque
                FROM homologacion_productos
            """).fetchall()
            self._catalog = {
                str(r["codigo_mp"]): HomologacionItem(
                    codigo_mp=str(r["codigo_mp"]),
                    itemcode_sap=r["itemcode_sap"],
                    descripcion_sap=r["descripcion_sap"],
                    factor_empaque=r["factor_empaque"] or 1.0,
                )
                for r in rows
            }
            self._loaded = True
        finally:
            conn.close()

    def _reload_sap_desc(self):
        conn = get_connection()
        try:
            rows = conn.execute(
                "SELECT itemcode_sap, descripcion_sap FROM sap_articulos WHERE activo=1"
            ).fetchall()
            self._sap_desc = {r["itemcode_sap"]: r["descripcion_sap"] for r in rows}
        finally:
            conn.close()

    # ------------------------------------------------------------------
    # Helpers privados para lectura de archivos
    # ------------------------------------------------------------------

    def _leer_maestra_excel(self, path: Path) -> Tuple[List[Tuple[str, str]], List[str]]:
        """Lee archivo xlsx buscando columnas de ItemCode y Descripción."""
        try:
            import openpyxl
        except ImportError:
            return [], ["openpyxl no instalado"]

        errores = []
        rows = []
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            ws = wb.active

            # Detectar índices de columna desde encabezados (fila 1)
            headers = [str(c).strip().upper() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            idx_code = _find_col(headers, ["ITEMCODE", "CODIGO", "COD SAP", "COD_SAP", "ITEM CODE", "ITEM_CODE"])
            idx_desc = _find_col(headers, ["DESCRIPCION", "DESCRIPTION", "NOMBRE", "DESCRIPCIÓN", "DESC"])

            if idx_code is None:
                return [], ["No se encontró columna de ItemCode en el archivo. Encabezados: " + ", ".join(headers)]
            if idx_desc is None:
                errores.append("No se encontró columna de descripción. Se usará vacío.")

            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[idx_code]).strip() if row[idx_code] else ""
                desc = str(row[idx_desc]).strip() if (idx_desc is not None and row[idx_desc]) else ""
                if code:
                    rows.append((code, desc))
            wb.close()
        except Exception as e:
            return [], [f"Error leyendo Excel: {e}"]

        return rows, errores

    def _leer_maestra_csv(self, path: Path) -> Tuple[List[Tuple[str, str]], List[str]]:
        """Lee archivo CSV buscando columnas de ItemCode y Descripción."""
        import csv
        errores = []
        rows = []
        try:
            # Detectar delimitador automáticamente
            with open(str(path), newline='', encoding='utf-8-sig') as f:
                sample = f.read(4096)
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            with open(str(path), newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f, dialect)
                headers = [h.strip().upper() for h in next(reader)]
                idx_code = _find_col(headers, ["ITEMCODE", "CODIGO", "COD SAP", "ITEM CODE"])
                idx_desc = _find_col(headers, ["DESCRIPCION", "DESCRIPTION", "NOMBRE", "DESC"])
                if idx_code is None:
                    return [], ["No se encontró columna de ItemCode en CSV."]
                for row in reader:
                    if len(row) <= idx_code:
                        continue
                    code = row[idx_code].strip()
                    desc = row[idx_desc].strip() if idx_desc is not None and len(row) > idx_desc else ""
                    if code:
                        rows.append((code, desc))
        except Exception as e:
            return [], [f"Error leyendo CSV: {e}"]
        return rows, errores


def _find_col(headers: List[str], candidates: List[str]) -> Optional[int]:
    """Busca el índice de la primera columna cuyo nombre coincide con algún candidato."""
    for i, h in enumerate(headers):
        for c in candidates:
            if c in h:
                return i
    return None


# Singleton global
_service: Optional[HomologacionService] = None


def get_homologacion_service() -> HomologacionService:
    global _service
    if _service is None:
        _service = HomologacionService()
    return _service
