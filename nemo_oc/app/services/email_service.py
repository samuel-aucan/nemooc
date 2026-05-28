"""
Servicio de notificaciones email para NemoOC.
Envia correos inmediatos por OC nueva y resumenes programados por cartera.
"""

import logging
import imaplib
import re
import smtplib
import ssl
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from html import escape
from pathlib import Path
from types import SimpleNamespace
from typing import Optional
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import openpyxl

from app.db import get_connection

logger = logging.getLogger(__name__)

COL_CARTERA = 0
COL_NOMBRE = 1
COL_CORREO = 2

HEADER_ALIASES = {
    "cartera": {"cartera", "holding"},
    "nombre": {"nombre", "vendedor"},
    "correo": {"correo", "email", "mail", "e-mail", "e_mail"},
    "activo": {"activo", "activa", "habilitado", "enabled"},
}

ACTIVE_TRUE_VALUES = {"1", "si", "sí", "true", "activo", "activa", "on", "yes", "y"}
ACTIVE_FALSE_VALUES = {"0", "no", "false", "inactivo", "off", "n", "disabled"}

SUMMARY_SCHEDULES = (
    {
        "key": "carryover_09",
        "hour": 9,
        "minute": 0,
        "subject_prefix": "Resumen arrastre OCs",
        "title": "Resumen de arrastre de OCs",
        "eyebrow": "NemoOC | Resumen automatico",
        "intro": "Resumen de arrastre con las ordenes recibidas ayer despues de las 17:00 para la cartera",
    },
    {
        "key": "daily_14",
        "hour": 14,
        "minute": 0,
        "subject_prefix": "Resumen OCs",
        "title": "Resumen diario de OCs",
        "eyebrow": "NemoOC | Resumen automatico",
        "intro": "Resumen consolidado de las ordenes recibidas hoy para la cartera",
    },
    {
        "key": "daily_17",
        "hour": 17,
        "minute": 0,
        "subject_prefix": "Resumen OCs",
        "title": "Resumen diario de OCs",
        "eyebrow": "NemoOC | Resumen automatico",
        "intro": "Resumen consolidado de las ordenes recibidas hoy para la cartera",
    },
)
DISPATCH_HEADER = "X-Nemo-Dispatch-Key"
SENT_MAILBOX_CANDIDATES = (
    "[Gmail]/Sent Mail",
    "[GoogleMail]/Sent Mail",
    "Sent",
    "Sent Items",
    "Sent Messages",
    "INBOX.Sent",
    "Enviados",
)
try:
    CHILE_TZ = ZoneInfo("America/Santiago")
except ZoneInfoNotFoundError:
    CHILE_TZ = None

_instance: Optional["EmailService"] = None


def get_email_service() -> "EmailService":
    global _instance
    if _instance is None:
        _instance = EmailService()
    return _instance


class EmailService:

    def __init__(self):
        self._vendedores: dict[str, list[dict]] = {}
        self._vendedores_list: list[dict] = []
        self._loaded = False
        self._sent_mailbox_cache: Optional[str] = None
        self._email_logo_cache: Optional[tuple[str, bytes, str]] = None
        self._email_logo_loaded = False

    # ------------------------------------------------------------------
    # Catalogo de vendedores
    # ------------------------------------------------------------------

    def cargar_correos(self, path: str | Path) -> tuple[bool, str]:
        """
        Lee CORREOS.xlsx y guarda el catalogo en SQLite.
        Formato soportado:
        - CARTERA | NOMBRE | CORREO
        - CARTERA | NOMBRE | CORREO | ACTIVO
        """
        try:
            path = Path(path)
            if not path.exists():
                return False, f"Archivo no encontrado: {path}"

            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                wb.close()
                return False, "El archivo de correos esta vacio."

            column_map, start_row = self._resolve_column_map(header_row)
            items: list[dict] = []

            for row in ws.iter_rows(min_row=start_row, values_only=True):
                cartera_raw = self._get_cell(row, column_map.get("cartera"))
                correo_raw = self._get_cell(row, column_map.get("correo"))
                if not cartera_raw or not correo_raw:
                    continue

                cartera = str(cartera_raw).strip().upper()
                nombre = str(self._get_cell(row, column_map.get("nombre")) or "").strip()
                correo = str(correo_raw).strip()
                activo = self._parse_activo(self._get_cell(row, column_map.get("activo")))

                if not cartera or not correo:
                    continue

                items.append({
                    "cartera": cartera,
                    "nombre": nombre,
                    "email": correo,
                    "activo": activo,
                })

            wb.close()

            if not items:
                return False, "No se encontraron vendedores validos en el archivo."

            self._guardar_vendedores(items, path.name)
            self.reload()

            total = len(self._vendedores_list)
            activos = sum(1 for item in self._vendedores_list if item["activo"])
            carteras = len({item["cartera"] for item in self._vendedores_list})
            return True, (
                f"{total} correo(s) cargados para {carteras} cartera(s). "
                f"Activos: {activos}"
            )

        except Exception as e:
            logger.error(f"Error cargando CORREOS.xlsx: {e}")
            return False, str(e)

    def reload(self) -> None:
        conn = get_connection()
        try:
            rows = conn.execute(
                """
                SELECT id, cartera, nombre, email, activo
                FROM vendedores_email
                ORDER BY UPPER(cartera), UPPER(nombre), UPPER(email), id
                """
            ).fetchall()
        finally:
            conn.close()

        mapping: dict[str, list[dict]] = defaultdict(list)
        vendedores_list: list[dict] = []
        for row in rows:
            item = {
                "id": int(row["id"]),
                "cartera": (row["cartera"] or "").strip().upper(),
                "nombre": (row["nombre"] or "").strip(),
                "email": (row["email"] or "").strip(),
                "activo": bool(row["activo"]),
            }
            vendedores_list.append(item)
            mapping[item["cartera"]].append(item)

        self._vendedores = dict(mapping)
        self._vendedores_list = vendedores_list
        self._loaded = True

    def count(self) -> int:
        if not self._loaded:
            self.reload()
        return len(self._vendedores_list)

    def listar_vendedores(self) -> list[dict]:
        if not self._loaded:
            self.reload()
        return [dict(item) for item in self._vendedores_list]

    def obtener_vendedor(self, vendedor_id: int) -> Optional[dict]:
        if not self._loaded:
            self.reload()
        for item in self._vendedores_list:
            if item["id"] == vendedor_id:
                return dict(item)
        return None

    def actualizar_vendedor_activo(self, vendedor_id: int, activo: bool) -> tuple[bool, str]:
        conn = get_connection()
        now = datetime.now().isoformat()
        try:
            result = conn.execute(
                """
                UPDATE vendedores_email
                SET activo = ?, updated_at = ?
                WHERE id = ?
                """,
                (1 if activo else 0, now, vendedor_id),
            )
            conn.commit()
            if result.rowcount == 0:
                return False, "Vendedor no encontrado."
        finally:
            conn.close()

        self.reload()
        item = self.obtener_vendedor(vendedor_id)
        if not item:
            return False, "Vendedor no encontrado."

        estado = "activado" if activo else "desactivado"
        etiqueta = item["nombre"] or item["email"]
        return True, f"{etiqueta} {estado}."

    def _guardar_vendedores(self, items: list[dict], origen_archivo: str) -> None:
        conn = get_connection()
        now = datetime.now().isoformat()
        try:
            conn.execute("DELETE FROM vendedores_email")
            for item in items:
                conn.execute(
                    """
                    INSERT INTO vendedores_email (
                        cartera, nombre, email, activo, origen_archivo,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item["cartera"],
                        item["nombre"],
                        item["email"],
                        1 if item["activo"] else 0,
                        origen_archivo,
                        now,
                        now,
                    ),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    # ------------------------------------------------------------------
    # Notificacion de OC nueva
    # ------------------------------------------------------------------

    def enviar_notificacion_oc(self, oc, cliente_info) -> bool:
        try:
            from app.config import load_config

            cfg = load_config()
            if not cfg.smtp_enabled:
                return False

            cartera = (cliente_info.cartera or "").strip().upper()
            destinatarios = self._destinatarios_cartera(cartera)
            if not destinatarios:
                logger.info(
                    f"Sin destinatarios activos para cartera '{cartera}' "
                    f"(OC {oc.codigo_oc})"
                )
                return False

            dispatch_key = self._build_notification_dispatch_key(cartera, oc.codigo_oc)
            dispatch_when = self._current_chile_time()
            if self._dispatch_ya_enviado(dispatch_key):
                return False
            if self._dispatch_existe_en_enviados(cfg, dispatch_key, dispatch_when):
                self._registrar_dispatch(
                    dispatch_key=dispatch_key,
                    dispatch_type="notification",
                    cartera=cartera,
                    scheduled_for=dispatch_when.isoformat(),
                )
                logger.info(
                    f"Notificacion OC {oc.codigo_oc} omitida: ya aparece en Enviados para cartera '{cartera}'."
                )
                return False

            lineas = self._obtener_lineas_oc(oc.codigo_oc)
            asunto = (
                f"Nueva OC {oc.tipo_oc or 'OC'} - {oc.codigo_oc} | "
                f"{oc.nombre_organismo or cliente_info.razon} | "
                f"{self._format_currency(oc.total)}"
            )
            cuerpo = self._construir_cuerpo_oc_nueva(oc, cliente_info, lineas)
            cuerpo_html = self._construir_cuerpo_oc_nueva_html(oc, cliente_info, lineas)
            ok = self._enviar_smtp(
                cfg=cfg,
                to_emails=destinatarios,
                cc_emails=self._build_cc_emails(cfg, destinatarios),
                asunto=asunto,
                cuerpo=cuerpo,
                cuerpo_html=cuerpo_html,
                dispatch_key=dispatch_key,
            )
            if ok:
                self._registrar_dispatch(
                    dispatch_key=dispatch_key,
                    dispatch_type="notification",
                    cartera=cartera,
                    scheduled_for=dispatch_when.isoformat(),
                )
            return ok

        except Exception as e:
            logger.warning(f"Error inesperado en enviar_notificacion_oc: {e}")
            return False

    def _construir_cuerpo_oc_nueva(self, oc, cliente_info, lineas: Optional[list] = None) -> str:
        portal_url = self._portal_url(oc.codigo_oc)
        organismo = oc.nombre_organismo or cliente_info.razon or ""
        tipo_oc = oc.tipo_oc or "OC"
        lineas = lineas or []

        detalle_lineas = [
            "",
            "DETALLE DE LINEAS",
            "----------------",
        ]
        if lineas:
            for linea in lineas:
                cantidad = self._format_quantity_with_unit(linea.cantidad, linea.unidad)
                cantidad_sap = self._format_linea_cantidad_sap(linea)
                producto = self._format_linea_producto(linea)
                detalle_linea = self._format_linea_detalle(linea)
                codigo_sap_sugerido = self._format_linea_codigo_sap_sugerido(linea)
                precio_sap = self._format_linea_precio_sap(linea)
                detalle_lineas.append(
                    f"{linea.correlativo}. {producto} | Cant: {cantidad} | Cant SAP: {cantidad_sap} | "
                    f"Precio neto: {self._format_currency(linea.precio_neto)} | "
                    f"Precio SAP: {precio_sap} | "
                    f"Total: {self._format_currency(linea.total)} | "
                    f"Codigo SAP sugerido: {codigo_sap_sugerido}"
                )
                if detalle_linea:
                    detalle_lineas.append(f"   Detalle: {detalle_linea}")
        else:
            detalle_lineas.append("No fue posible cargar el detalle de productos de esta OC.")

        return (
            f"Se ha recibido una nueva {tipo_oc} en NemoOC.\n\n"
            f"DATOS DE LA OC\n"
            f"--------------\n"
            f"Codigo OC   : {oc.codigo_oc}\n"
            f"Tipo        : {tipo_oc}\n"
            f"Organismo   : {organismo}\n"
            f"Cartera     : {cliente_info.cartera}\n"
            f"Region      : {cliente_info.region_nombre}\n"
            f"Monto Total : {self._format_currency(oc.total)}\n"
            f"Estado SAP  : {self._format_estado_sap(oc.estado_interno, oc.fecha_ingreso)}\n"
            f"Fecha envio : {oc.fecha_envio[:10] if oc.fecha_envio else '-'}\n"
            f"Lineas      : {len(lineas) or getattr(oc, 'cantidad_lineas', 0) or '-'}\n"
            f"{chr(10).join(detalle_lineas)}\n\n"
            f"Ver en portal:\n{portal_url}\n\n"
            f"---\n"
            f"Este mensaje fue generado automaticamente por NemoOC.\n"
        )

    def _construir_cuerpo_oc_nueva_html(self, oc, cliente_info, lineas: Optional[list] = None) -> str:
        portal_url = self._portal_url(oc.codigo_oc)
        organismo = oc.nombre_organismo or cliente_info.razon or ""
        tipo_oc = oc.tipo_oc or "OC"
        lineas = lineas or []
        cantidad_lineas = len(lineas) or getattr(oc, "cantidad_lineas", 0) or 0

        detalle_html = self._render_key_value_table_html([
            ("Codigo OC", oc.codigo_oc),
            ("Tipo", tipo_oc),
            ("Organismo", organismo),
            ("Cartera", cliente_info.cartera or "-"),
            ("Region", cliente_info.region_nombre or "-"),
            ("Monto total", self._format_currency(oc.total)),
            ("Estado SAP", self._format_estado_sap(oc.estado_interno, oc.fecha_ingreso)),
            ("Fecha envio", self._format_date_only(oc.fecha_envio)),
            ("Lineas", str(cantidad_lineas or "-")),
        ])

        metricas_html = self._render_metric_cards_html([
            ("Monto", self._format_currency(oc.total)),
            ("Cartera", cliente_info.cartera or "-"),
            ("Estado SAP", self._format_estado_sap(oc.estado_interno, oc.fecha_ingreso)),
            ("Lineas", str(cantidad_lineas or "-")),
        ])

        if lineas:
            detalle_lineas_html = self._render_oc_lines_table_html(lineas)
        else:
            detalle_lineas_html = (
                "<div style=\"margin-top:16px;padding:14px 16px;border:1px dashed #d8e1ee;"
                "border-radius:14px;background:#fbfdff;font-size:13px;line-height:1.6;color:#6b7280;\">"
                "No fue posible cargar el detalle de lineas de esta OC al momento de generar el correo."
                "</div>"
            )

        cuerpo_html = (
            f"{metricas_html}"
            f"<p style=\"margin:0 0 18px 0;font-size:14px;line-height:1.6;color:#4b5563;\">"
            f"Se detecto una nueva <strong>{self._html_escape(tipo_oc)}</strong> asociada a la cartera "
            f"<strong>{self._html_escape(cliente_info.cartera or '-')}</strong>. "
            "Debajo va el detalle consolidado para que puedan gestionarla rapido."
            f"</p>"
            f"{detalle_html}"
            "<div style=\"margin:22px 0 10px 0;font-size:16px;line-height:1.4;font-weight:700;color:#172033;\">"
            "Detalle de productos"
            "</div>"
            "<p style=\"margin:0 0 14px 0;font-size:13px;line-height:1.6;color:#6b7280;\">"
            "Se listan las lineas detectadas en la OC para que puedan revisar productos, cantidades y montos "
            "sin salir del correo."
            "</p>"
            f"{detalle_lineas_html}"
            f"{self._render_button_html(portal_url, 'Ver OC en Mercado Publico')}"
        )
        return self._build_email_layout_html(
            eyebrow="NemoOC | Notificacion automatica",
            title=f"Nueva {tipo_oc} recibida",
            subtitle=f"{oc.codigo_oc} · {organismo or 'Sin organismo'}",
            body_html=cuerpo_html,
        )

    # ------------------------------------------------------------------
    # Resumenes programados
    # ------------------------------------------------------------------

    def enviar_resumenes_programados(self, now: Optional[datetime] = None) -> int:
        from app.config import load_config

        cfg = load_config()
        if not cfg.smtp_enabled:
            return 0

        if CHILE_TZ is None:
            now_chile = now or datetime.now()
        else:
            now_chile = now.astimezone(CHILE_TZ) if now else datetime.now(CHILE_TZ)
        enviados = 0

        for schedule in SUMMARY_SCHEDULES:
            slot_dt = now_chile.replace(
                hour=schedule["hour"],
                minute=schedule["minute"],
                second=0,
                microsecond=0,
            )
            if now_chile < slot_dt:
                continue
            enviados += self._enviar_resumenes_slot(cfg, slot_dt, schedule)

        return enviados

    def _enviar_resumenes_slot(self, cfg, slot_dt: datetime, schedule: dict) -> int:
        summary_ctx = self._build_summary_context(slot_dt, schedule)
        agrupadas = self._resumen_ocs_por_cartera(
            window_start_dt=summary_ctx["window_start_dt"],
            window_end_dt=summary_ctx["window_end_dt"],
        )

        enviadas = 0
        for cartera, rows in agrupadas.items():
            dispatch_key = self._build_summary_dispatch_key(
                cartera=cartera,
                slot_dt=slot_dt,
                schedule_key=summary_ctx["schedule_key"],
            )
            if self._dispatch_ya_enviado(dispatch_key):
                continue
            if self._dispatch_existe_en_enviados(cfg, dispatch_key, slot_dt):
                self._registrar_dispatch(
                    dispatch_key=dispatch_key,
                    dispatch_type="summary",
                    cartera=cartera,
                    scheduled_for=slot_dt.isoformat(),
                )
                logger.info(
                    "Resumen %s omitido para cartera '%s': ya existe en Enviados.",
                    summary_ctx["log_label"],
                    cartera,
                )
                continue

            destinatarios = self._destinatarios_cartera(cartera)
            if not destinatarios:
                logger.info(
                    "Resumen %s omitido para cartera '%s': sin destinatarios activos.",
                    summary_ctx["log_label"],
                    cartera,
                )
                continue

            asunto = (
                f"{summary_ctx['subject_prefix']} {cartera} - "
                f"{slot_dt.strftime('%d/%m/%Y %H:%M')} CL"
            )
            cuerpo = self._construir_cuerpo_resumen(
                cartera=cartera,
                slot_dt=slot_dt,
                rows=rows,
                summary_ctx=summary_ctx,
            )
            cuerpo_html = self._construir_cuerpo_resumen_html(
                cartera=cartera,
                slot_dt=slot_dt,
                rows=rows,
                summary_ctx=summary_ctx,
            )
            ok = self._enviar_smtp(
                cfg=cfg,
                to_emails=destinatarios,
                cc_emails=self._build_cc_emails(cfg, destinatarios),
                asunto=asunto,
                cuerpo=cuerpo,
                cuerpo_html=cuerpo_html,
                dispatch_key=dispatch_key,
            )
            if ok:
                self._registrar_dispatch(
                    dispatch_key=dispatch_key,
                    dispatch_type="summary",
                    cartera=cartera,
                    scheduled_for=slot_dt.isoformat(),
                )
                enviadas += 1

        return enviadas

    def _resumen_ocs_por_cartera(
        self,
        window_start_dt: datetime,
        window_end_dt: datetime,
    ) -> dict[str, list[dict]]:
        window_start = self._format_db_datetime(window_start_dt)
        window_end = self._format_db_datetime(window_end_dt)
        conn = get_connection()
        try:
            rows = conn.execute(
                """
                SELECT
                    UPPER(TRIM(c.cartera)) AS cartera,
                    o.codigo_oc,
                    COALESCE(NULLIF(TRIM(o.tipo_oc), ''), 'CM') AS tipo_oc,
                    COALESCE(NULLIF(TRIM(o.nombre_organismo), ''), 'Sin organismo') AS nombre_organismo,
                    COALESCE(o.total, o.total_neto, 0) AS total,
                    COALESCE(NULLIF(TRIM(o.estado_interno), ''), 'Nueva') AS estado_interno,
                    COALESCE(o.fecha_ingreso, '') AS fecha_ingreso,
                    COALESCE(o.estado_mp, '') AS estado_mp,
                    COALESCE(o.created_at, '') AS created_at
                FROM oc_cabecera o
                INNER JOIN cartera_clientes c
                    ON TRIM(c.cod_cliente) = TRIM(o.cliente_sap_sugerido)
                WHERE COALESCE(TRIM(c.cartera), '') != ''
                  AND COALESCE(o.created_at, '') >= ?
                  AND COALESCE(o.created_at, '') < ?
                ORDER BY UPPER(TRIM(c.cartera)), COALESCE(o.created_at, ''), o.codigo_oc
                """,
                (window_start, window_end),
            ).fetchall()
        finally:
            conn.close()

        agrupadas: dict[str, list[dict]] = defaultdict(list)
        for row in rows:
            cartera = (row["cartera"] or "").strip().upper()
            if not cartera:
                continue
            agrupadas[cartera].append(dict(row))
        return dict(agrupadas)

    def _construir_cuerpo_resumen(
        self,
        cartera: str,
        slot_dt: datetime,
        rows: list[dict],
        summary_ctx: dict,
    ) -> str:
        ingresadas = sum(
            1 for row in rows
            if (row.get("estado_interno") or "").strip().lower() == "ingresada"
        )
        pendientes = len(rows) - ingresadas

        lineas = [
            summary_ctx["title"],
            "",
            f"Fecha       : {slot_dt.strftime('%d/%m/%Y')}",
            f"Corte       : {slot_dt.strftime('%H:%M')} hora Chile",
            f"Ventana     : {summary_ctx['window_label']}",
            f"Cartera     : {cartera}",
            f"Total OCs   : {len(rows)}",
            f"Ingresadas  : {ingresadas}",
            f"Pendientes  : {pendientes}",
            "",
            "DETALLE",
            "-------",
        ]

        for row in rows:
            llegada = self._format_datetime_short(row.get("created_at"))
            estado_sap = self._format_estado_sap(
                row.get("estado_interno", ""),
                row.get("fecha_ingreso", ""),
            )
            estado_mp = row.get("estado_mp") or "-"
            lineas.append(
                f"- {row['codigo_oc']} | {row['tipo_oc']} | {row['nombre_organismo']} | "
                f"{self._format_currency(row.get('total'))} | "
                f"llego {llegada} | SAP: {estado_sap} | MP: {estado_mp}"
            )

        lineas.extend([
            "",
            "---",
            "Este mensaje fue generado automaticamente por NemoOC.",
        ])
        return "\n".join(lineas)

    def _construir_cuerpo_resumen_html(
        self,
        cartera: str,
        slot_dt: datetime,
        rows: list[dict],
        summary_ctx: dict,
    ) -> str:
        ingresadas = sum(
            1 for row in rows
            if (row.get("estado_interno") or "").strip().lower() == "ingresada"
        )
        pendientes = len(rows) - ingresadas

        metricas_html = self._render_metric_cards_html([
            ("Total OCs", str(len(rows))),
            ("Ingresadas", str(ingresadas)),
            ("Pendientes", str(pendientes)),
            ("Corte", f"{slot_dt.strftime('%H:%M')} CL"),
        ])

        tabla_html = self._render_summary_table_html(rows)
        cuerpo_html = (
            f"{metricas_html}"
            f"{self._render_key_value_table_html([('Ventana', summary_ctx['window_label'])])}"
            "<div style=\"height:14px;\"></div>"
            f"<p style=\"margin:0 0 18px 0;font-size:14px;line-height:1.6;color:#4b5563;\">"
            f"{self._html_escape(summary_ctx['intro'])} "
            f"<strong>{self._html_escape(cartera)}</strong>. "
            "La columna SAP indica si ya fueron ingresadas o si siguen pendientes."
            f"</p>"
            f"{tabla_html}"
        )

        return self._build_email_layout_html(
            eyebrow=summary_ctx["eyebrow"],
            title=f"{summary_ctx['title']} - {cartera}",
            subtitle=(
                f"{slot_dt.strftime('%d/%m/%Y %H:%M')} hora Chile | "
                f"Ventana {summary_ctx['window_label']}"
            ),
            body_html=cuerpo_html,
        )

    def _build_summary_context(self, slot_dt: datetime, schedule: dict) -> dict:
        schedule_key = str(schedule.get("key") or "").strip()
        subject_prefix = str(schedule.get("subject_prefix") or "Resumen OCs").strip()
        title = str(schedule.get("title") or "Resumen de OCs").strip()
        eyebrow = str(schedule.get("eyebrow") or "NemoOC | Resumen automatico").strip()
        intro = str(schedule.get("intro") or "Resumen de las ordenes recibidas para la cartera").strip()

        if schedule_key == "carryover_09":
            midnight = slot_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            window_start_dt = midnight - timedelta(hours=7)
            window_end_dt = midnight
            log_label = "arrastre 09:00"
        else:
            day_start_dt = slot_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            window_start_dt = day_start_dt
            window_end_dt = slot_dt + timedelta(seconds=1)
            log_label = slot_dt.strftime("%H:%M")

        return {
            "schedule_key": schedule_key,
            "subject_prefix": subject_prefix,
            "title": title,
            "eyebrow": eyebrow,
            "intro": intro,
            "window_start_dt": window_start_dt,
            "window_end_dt": window_end_dt,
            "window_label": self._format_summary_window(window_start_dt, window_end_dt),
            "log_label": log_label,
        }

    def _format_summary_window(self, window_start_dt: datetime, window_end_dt: datetime) -> str:
        display_end_dt = window_end_dt - timedelta(seconds=1)
        if display_end_dt < window_start_dt:
            display_end_dt = window_start_dt
        return (
            f"{window_start_dt.strftime('%d/%m %H:%M')} - "
            f"{display_end_dt.strftime('%d/%m %H:%M')} CL"
        )

    def _format_db_datetime(self, value: datetime) -> str:
        if value.tzinfo is not None:
            if CHILE_TZ is not None:
                value = value.astimezone(CHILE_TZ)
            else:
                value = value.astimezone().replace(tzinfo=None)
        if value.tzinfo is not None:
            value = value.replace(tzinfo=None)
        return value.isoformat(timespec="seconds")

    # ------------------------------------------------------------------
    # Dispatch log
    # ------------------------------------------------------------------

    def _build_notification_dispatch_key(self, cartera: str, codigo_oc: str) -> str:
        return f"notification::{(cartera or '').strip().upper()}::{str(codigo_oc or '').strip()}"

    def _build_summary_dispatch_key(self, cartera: str, slot_dt: datetime, schedule_key: str) -> str:
        return (
            f"summary::{schedule_key}::{slot_dt.strftime('%Y-%m-%dT%H:%M')}::"
            f"{(cartera or '').strip().upper()}"
        )

    def _dispatch_ya_enviado(self, dispatch_key: str) -> bool:
        conn = get_connection()
        try:
            row = conn.execute(
                "SELECT 1 FROM notification_dispatch_log WHERE dispatch_key = ?",
                (dispatch_key,),
            ).fetchone()
            return row is not None
        finally:
            conn.close()

    def _registrar_dispatch(
        self,
        *,
        dispatch_key: str,
        dispatch_type: str,
        cartera: str = "",
        scheduled_for: str = "",
    ) -> None:
        conn = get_connection()
        now = datetime.now().isoformat()
        try:
            conn.execute(
                """
                INSERT OR IGNORE INTO notification_dispatch_log (
                    dispatch_key, dispatch_type, cartera, scheduled_for,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (dispatch_key, dispatch_type, cartera, scheduled_for, now, now),
            )
            conn.commit()
        finally:
            conn.close()

    def _dispatch_existe_en_enviados(self, cfg, dispatch_key: str, when: datetime) -> bool:
        if not getattr(cfg, "smtp_user", "") or not getattr(cfg, "smtp_password", ""):
            return False

        imap = None
        try:
            imap = imaplib.IMAP4_SSL(cfg.imap_server, cfg.imap_port)
            imap.login(cfg.smtp_user, cfg.smtp_password)
            mailbox = self._resolve_sent_mailbox(imap)
            status, _ = imap.select(mailbox, readonly=True)
            if status != "OK":
                return False

            criteria = self._format_imap_date(when)
            if self._imap_search_has_results(
                imap,
                "SINCE",
                criteria,
                "HEADER",
                DISPATCH_HEADER,
                dispatch_key,
            ):
                return True

            if self._imap_search_has_results(
                imap,
                "SINCE",
                criteria,
                "TEXT",
                dispatch_key,
            ):
                return True

            for legacy_term in self._legacy_dispatch_search_terms(dispatch_key):
                if self._imap_search_has_results(
                    imap,
                    "SINCE",
                    criteria,
                    "TEXT",
                    legacy_term,
                ):
                    logger.info(
                        "Deduplicacion por Enviados: '%s' ya aparece hoy con termino legacy '%s'.",
                        dispatch_key,
                        legacy_term,
                    )
                    return True
            return False
        except Exception as e:
            logger.warning(f"No fue posible revisar Enviados para deduplicar '{dispatch_key}': {e}")
            return False
        finally:
            if imap is not None:
                try:
                    imap.logout()
                except Exception:
                    pass

    def _imap_search_has_results(self, imap, *criteria: str) -> bool:
        status, data = imap.search(None, *criteria)
        return bool(status == "OK" and data and data[0].strip())

    def _legacy_dispatch_search_terms(self, dispatch_key: str) -> list[str]:
        key = (dispatch_key or "").strip()
        if not key.casefold().startswith("notification::"):
            return []

        parts = key.split("::")
        if len(parts) < 3:
            return []

        codigo_oc = parts[-1].strip()
        if not codigo_oc:
            return []

        return [codigo_oc]

    def _resolve_sent_mailbox(self, imap) -> str:
        if self._sent_mailbox_cache:
            return self._sent_mailbox_cache

        status, mailboxes = imap.list()
        if status == "OK":
            for raw in mailboxes or []:
                line = raw.decode("utf-8", errors="ignore")
                mailbox = self._parse_imap_mailbox_name(line)
                if mailbox and "\\sent" in line.casefold():
                    self._sent_mailbox_cache = mailbox
                    return mailbox

            parsed = [
                self._parse_imap_mailbox_name(raw.decode("utf-8", errors="ignore"))
                for raw in (mailboxes or [])
            ]
            normalized = {name.casefold(): name for name in parsed if name}
            for candidate in SENT_MAILBOX_CANDIDATES:
                resolved = normalized.get(candidate.casefold())
                if resolved:
                    self._sent_mailbox_cache = resolved
                    return resolved

        self._sent_mailbox_cache = "[Gmail]/Sent Mail"
        return self._sent_mailbox_cache

    def _parse_imap_mailbox_name(self, line: str) -> str:
        match = re.match(r'.*\s(?P<name>".*"|[^"]+)$', line.strip())
        if not match:
            return ""
        mailbox = match.group("name").strip()
        if mailbox.startswith('"') and mailbox.endswith('"'):
            mailbox = mailbox[1:-1]
        return mailbox

    def _format_imap_date(self, value: datetime) -> str:
        months = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
        return f"{value.day:02d}-{months[value.month - 1]}-{value.year}"

    def _current_chile_time(self) -> datetime:
        if CHILE_TZ is None:
            return datetime.now()
        return datetime.now(CHILE_TZ)

    # ------------------------------------------------------------------
    # Envio SMTP
    # ------------------------------------------------------------------

    def _enviar_smtp(
        self,
        cfg,
        to_emails: list[str],
        cc_emails: list[str],
        asunto: str,
        cuerpo: str,
        cuerpo_html: str = "",
        dispatch_key: str = "",
    ) -> bool:
        try:
            to_emails = self._unique_emails(to_emails)
            to_lookup = {email.casefold() for email in to_emails}
            cc_emails = self._unique_emails(
                [mail for mail in cc_emails if mail.casefold() not in to_lookup]
            )
            recipients = to_emails + cc_emails
            if not recipients:
                return False

            logo = self._get_email_logo() if cuerpo_html else None
            if cuerpo_html and logo:
                msg = MIMEMultipart("related")
                alternative = MIMEMultipart("alternative")
                msg.attach(alternative)
            else:
                msg = MIMEMultipart("alternative")
                alternative = msg

            msg["Subject"] = asunto
            msg["From"] = cfg.smtp_user
            msg["To"] = ", ".join(to_emails)
            if cc_emails:
                msg["Cc"] = ", ".join(cc_emails)
            if dispatch_key:
                msg[DISPATCH_HEADER] = dispatch_key
            alternative.attach(MIMEText(cuerpo, "plain", "utf-8"))
            if cuerpo_html:
                alternative.attach(MIMEText(cuerpo_html, "html", "utf-8"))
            if logo:
                content_id, logo_bytes, subtype = logo
                image_part = MIMEImage(logo_bytes, _subtype=subtype)
                image_part.add_header("Content-ID", f"<{content_id}>")
                image_part.add_header("Content-Disposition", "inline", filename=f"{content_id}.{subtype}")
                msg.attach(image_part)

            context = ssl.create_default_context()
            with smtplib.SMTP(cfg.smtp_host, cfg.smtp_port, timeout=15) as server:
                server.ehlo()
                server.starttls(context=context)
                server.ehlo()
                server.login(cfg.smtp_user, cfg.smtp_password)
                server.sendmail(cfg.smtp_user, recipients, msg.as_string())

            logger.info(f"Email enviado a {', '.join(recipients)}: {asunto[:80]}")
            return True

        except smtplib.SMTPAuthenticationError:
            logger.warning(
                f"SMTP auth fallida para {cfg.smtp_user}. "
                "Verifique usuario/contraseña en Configuracion."
            )
            return False
        except smtplib.SMTPException as e:
            logger.warning(f"SMTP error enviando '{asunto[:60]}': {e}")
            return False
        except OSError as e:
            logger.warning(f"Error de red SMTP ({cfg.smtp_host}:{cfg.smtp_port}): {e}")
            return False
        except Exception as e:
            logger.warning(f"Error inesperado en _enviar_smtp: {e}")
            return False

    def enviar_prueba(self, cfg, to_email: str = "") -> tuple[bool, str]:
        if not cfg.smtp_user:
            return False, "Configure el usuario/email primero."

        destino = to_email.strip() or cfg.smtp_user
        oc_demo, cliente_demo, lineas_demo = self._build_demo_oc_email_context()
        asunto = (
            f"PRUEBA DETALLE HTML - Nueva OC {oc_demo.tipo_oc} - "
            f"{oc_demo.codigo_oc}"
        )
        cuerpo = self._construir_cuerpo_prueba_texto(
            cfg,
            destino,
            oc_demo,
            cliente_demo,
            lineas_demo,
        )
        cuerpo_html = self._construir_cuerpo_prueba_html(
            cfg,
            destino,
            oc_demo,
            cliente_demo,
            lineas_demo,
        )
        ok = self._enviar_smtp(
            cfg=cfg,
            to_emails=[destino],
            cc_emails=[],
            asunto=asunto,
            cuerpo=cuerpo,
            cuerpo_html=cuerpo_html,
        )
        if ok:
            return True, f"Email de prueba enviado a {destino}"
        return False, "Envio fallido. Revise credenciales y conectividad."

    def _construir_cuerpo_prueba_texto(self, cfg, destino: str, oc, cliente_info, lineas: list) -> str:
        ahora = datetime.now(CHILE_TZ) if CHILE_TZ else datetime.now()
        detalle_oc = self._construir_cuerpo_oc_nueva(oc, cliente_info, lineas)
        return (
            "Este es un correo de prueba de NemoOC usando una OC ficticia.\n"
            "El objetivo es validar como veran vendedores y usuarios el formato final.\n\n"
            "RESUMEN\n"
            "-------\n"
            f"Destino       : {destino}\n"
            f"Remitente     : {cfg.smtp_user}\n"
            f"Servidor SMTP : {cfg.smtp_host}:{cfg.smtp_port}\n"
            f"Hora envio    : {ahora.strftime('%d/%m/%Y %H:%M')}\n"
            f"OC prueba     : {oc.codigo_oc}\n\n"
            f"{detalle_oc}"
        )

    def _construir_cuerpo_prueba_html(self, cfg, destino: str, oc, cliente_info, lineas: list) -> str:
        ahora = datetime.now(CHILE_TZ) if CHILE_TZ else datetime.now()
        detalle_oc = self._construir_cuerpo_oc_nueva_html(oc, cliente_info, lineas)
        cuerpo_html = (
            "<p style=\"margin:0 0 18px 0;font-size:14px;line-height:1.6;color:#4b5563;\">"
            "Esta es una validacion del canal SMTP de NemoOC usando una OC ficticia completa. "
            "Incluye cabecera, detalle de lineas y codigo SAP sugerido para simular el correo real."
            "</p>"
            + self._render_key_value_table_html([
                ("Destino", destino),
                ("Remitente", cfg.smtp_user),
                ("Servidor SMTP", f"{cfg.smtp_host}:{cfg.smtp_port}"),
                ("Hora envio", ahora.strftime("%d/%m/%Y %H:%M")),
                ("OC prueba", oc.codigo_oc),
            ])
            + "<div style=\"height:18px;\"></div>"
            + detalle_oc
        )
        return self._build_email_layout_html(
            eyebrow="NemoOC | Prueba de OC",
            title=f"Prueba de notificacion - {oc.tipo_oc}",
            subtitle=f"{destino} · {oc.codigo_oc}",
            body_html=cuerpo_html,
        )

    def _build_demo_oc_email_context(self) -> tuple[object, object, list]:
        fecha_envio = datetime.now(CHILE_TZ) if CHILE_TZ else datetime.now()
        fecha_envio_iso = fecha_envio.replace(microsecond=0).isoformat()

        oc_demo = SimpleNamespace(
            codigo_oc="PRUEBA-OC-DETALLE-20260420-001",
            tipo_oc="CM",
            nombre_organismo="Hospital Clinico de Prueba",
            total=1840000,
            estado_interno="Nueva",
            fecha_ingreso="",
            fecha_envio=fecha_envio_iso,
            cantidad_lineas=3,
        )

        cliente_demo = SimpleNamespace(
            razon="Hospital Clinico de Prueba",
            cartera="ATEL",
            region_nombre="Metropolitana",
        )

        lineas_demo = [
            SimpleNamespace(
                correlativo=1,
                producto="Guante quirurgico esteril",
                especificacion_comprador="Talla M, caja de 100 unidades",
                especificacion_proveedor="",
                codigo_mp="MP-TEST-001",
                cantidad=12,
                unidad="cajas",
                precio_neto=45000,
                total=540000,
                itemcode_sap="SAP-GUA-001",
                descripcion_sap="GUANTE QUIRURGICO ESTERIL M",
            ),
            SimpleNamespace(
                correlativo=2,
                producto="Mascarilla N95",
                especificacion_comprador="Caja de 20 unidades, certificacion clinica",
                especificacion_proveedor="",
                codigo_mp="MP-TEST-002",
                cantidad=26,
                unidad="cajas",
                precio_neto=50000,
                total=1300000,
                itemcode_sap="SAP-MSK-095",
                descripcion_sap="MASCARILLA N95 CLINICA",
            ),
            SimpleNamespace(
                correlativo=3,
                producto="Pechera desechable manga larga",
                especificacion_comprador="Uso clinico, talla universal",
                especificacion_proveedor="",
                codigo_mp="MP-TEST-003",
                cantidad=8,
                unidad="bolsas",
                precio_neto=0,
                total=0,
                itemcode_sap="",
                descripcion_sap="",
            ),
        ]
        return oc_demo, cliente_demo, lineas_demo

    def _build_email_layout_html(
        self,
        *,
        eyebrow: str,
        title: str,
        subtitle: str,
        body_html: str,
    ) -> str:
        logo_html = self._render_email_logo_html()
        return (
            "<!DOCTYPE html>"
            "<html lang=\"es\">"
            "<head>"
            "<meta charset=\"utf-8\">"
            "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">"
            "<style>"
            "body{margin:0;padding:0;background:#eef3f9;font-family:Segoe UI,Arial,sans-serif;color:#172033;}"
            ".mobile-only{display:none;max-height:0;overflow:hidden;}"
            "@media only screen and (max-width:640px){"
            ".email-shell{padding:12px !important;}"
            ".email-card{border-radius:16px !important;}"
            ".email-header{padding:22px 18px !important;}"
            ".email-body{padding:20px 16px !important;}"
            ".email-footer{padding:16px 16px !important;}"
            ".email-title{font-size:22px !important;line-height:1.25 !important;}"
            ".email-subtitle{font-size:13px !important;line-height:1.45 !important;}"
            ".metric-col{display:block !important;width:100% !important;padding:0 0 10px 0 !important;}"
            ".metric-col table{width:100% !important;}"
            ".kv-label,.kv-value{display:block !important;width:100% !important;box-sizing:border-box;}"
            ".kv-label{padding:12px 14px 4px 14px !important;border-bottom:0 !important;}"
            ".kv-value{padding:0 14px 12px 14px !important;border-bottom:1px solid #e5eaf3 !important;}"
            ".desktop-only{display:none !important;max-height:0 !important;overflow:hidden !important;mso-hide:all !important;}"
            ".mobile-only{display:block !important;max-height:none !important;overflow:visible !important;}"
            ".mobile-summary-card{display:block !important;}"
            ".mobile-summary-title{font-size:16px !important;line-height:1.35 !important;}"
            ".mobile-summary-meta{font-size:13px !important;line-height:1.5 !important;}"
            "}"
            "</style>"
            "</head>"
            "<body style=\"margin:0;padding:0;background:#eef3f9;font-family:Segoe UI,Arial,sans-serif;color:#172033;\">"
            "<table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" "
            "class=\"email-shell\" style=\"background:#eef3f9;padding:24px 12px;\">"
            "<tr><td align=\"center\">"
            "<table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" "
            "class=\"email-card\" style=\"max-width:760px;background:#ffffff;border:1px solid #d8e1ee;border-radius:18px;overflow:hidden;\">"
            "<tr><td class=\"email-header\" style=\"padding:28px 32px;background:#0f2742;\">"
            f"{logo_html}"
            f"<div style=\"font-size:11px;letter-spacing:0.18em;text-transform:uppercase;color:#9fc1ff;"
            "font-weight:700;margin-bottom:10px;\">"
            f"{self._html_escape(eyebrow)}</div>"
            f"<div class=\"email-title\" style=\"font-size:28px;line-height:1.2;font-weight:700;color:#ffffff;\">"
            f"{self._html_escape(title)}</div>"
            f"<div class=\"email-subtitle\" style=\"font-size:14px;line-height:1.5;color:#c8d7ea;margin-top:8px;\">"
            f"{self._html_escape(subtitle)}</div>"
            "</td></tr>"
            "<tr><td class=\"email-body\" style=\"padding:28px 32px;\">"
            f"{body_html}"
            "</td></tr>"
            "<tr><td class=\"email-footer\" style=\"padding:18px 32px;background:#f8fbff;border-top:1px solid #d8e1ee;"
            "font-size:12px;line-height:1.6;color:#6b7280;\">"
            "Mensaje generado automaticamente por NemoOC. "
            "Si necesitas cambios en destinatarios o automatizaciones, revisalos desde la configuracion web."
            "</td></tr>"
            "</table>"
            "</td></tr></table>"
            "</body></html>"
        )

    def _render_email_logo_html(self) -> str:
        logo = self._get_email_logo()
        if not logo:
            return ""

        content_id, _, _ = logo
        return (
            "<div style=\"margin:0 0 18px 0;\">"
            "<div style=\"display:inline-block;padding:8px;border-radius:18px;"
            "background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.12);\">"
            f"<img src=\"cid:{content_id}\" alt=\"Nemono\" width=\"68\" "
            "style=\"display:block;width:68px;max-width:68px;height:auto;border:0;outline:none;text-decoration:none;\">"
            "</div>"
            "</div>"
        )

    def _get_email_logo(self) -> Optional[tuple[str, bytes, str]]:
        if self._email_logo_loaded:
            return self._email_logo_cache

        self._email_logo_loaded = True
        for candidate in self._iter_email_logo_candidates():
            try:
                if not candidate.exists() or not candidate.is_file():
                    continue

                suffix = candidate.suffix.lower()
                subtype = "png" if suffix == ".png" else "jpeg" if suffix in {".jpg", ".jpeg"} else ""
                if not subtype:
                    continue

                self._email_logo_cache = ("nemono-logo", candidate.read_bytes(), subtype)
                return self._email_logo_cache
            except Exception as e:
                logger.debug(f"No se pudo cargar logo de correo desde {candidate}: {e}")

        return None

    def _iter_email_logo_candidates(self) -> list[Path]:
        workspace_root = Path(__file__).resolve().parents[3]
        candidates = [
            workspace_root / "nemo_oc_web" / "frontend" / "public" / "branding" / "nemono-favicon.png",
            workspace_root / "nemo_oc" / "assets" / "mono.png",
        ]

        meipass = getattr(sys, "_MEIPASS", "")
        if meipass:
            meipass_root = Path(meipass)
            candidates = [
                meipass_root / "branding" / "nemono-favicon.png",
                meipass_root / "nemo_oc" / "assets" / "mono.png",
                *candidates,
            ]

        unique_candidates: list[Path] = []
        seen: set[str] = set()
        for candidate in candidates:
            normalized = str(candidate.resolve(strict=False)).casefold()
            if normalized in seen:
                continue
            seen.add(normalized)
            unique_candidates.append(candidate)
        return unique_candidates

    def _render_key_value_table_html(self, rows: list[tuple[str, str]]) -> str:
        rendered_rows = []
        for label, value in rows:
            rendered_rows.append(
                "<tr>"
                "<td class=\"kv-label\" style=\"width:34%;padding:12px 14px;border-bottom:1px solid #e5eaf3;"
                "font-size:12px;letter-spacing:0.08em;text-transform:uppercase;color:#6b7280;\">"
                f"{self._html_escape(label)}</td>"
                "<td class=\"kv-value\" style=\"padding:12px 14px;border-bottom:1px solid #e5eaf3;"
                "font-size:14px;line-height:1.5;color:#172033;font-weight:600;\">"
                f"{self._html_escape(value)}</td>"
                "</tr>"
            )
        return (
            "<table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" "
            "style=\"border:1px solid #d8e1ee;border-radius:14px;overflow:hidden;background:#fbfdff;"
            "border-collapse:separate;border-spacing:0;\">"
            f"{''.join(rendered_rows)}"
            "</table>"
        )

    def _render_metric_cards_html(self, metrics: list[tuple[str, str]]) -> str:
        if not metrics:
            return ""
        cells = []
        for label, value in metrics:
            cells.append(
                "<td class=\"metric-col\" style=\"padding:0 10px 12px 0;vertical-align:top;\">"
                "<table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" "
                "style=\"background:#f5f9ff;border:1px solid #d8e1ee;border-radius:14px;\">"
                "<tr><td style=\"padding:14px 16px;\">"
                "<div style=\"font-size:11px;letter-spacing:0.14em;text-transform:uppercase;color:#6b7280;\">"
                f"{self._html_escape(label)}</div>"
                "<div style=\"margin-top:8px;font-size:20px;line-height:1.2;font-weight:700;color:#0f2742;\">"
                f"{self._html_escape(value)}</div>"
                "</td></tr></table>"
                "</td>"
            )
        return (
            "<table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" "
            "style=\"margin:0 0 18px 0;\"><tr>"
            f"{''.join(cells)}"
            "</tr></table>"
        )

    def _render_oc_lines_table_html(self, lineas: list) -> str:
        table_rows: list[str] = []
        for index, linea in enumerate(lineas):
            background = "#ffffff" if index % 2 == 0 else "#f8fbff"
            producto = self._format_linea_producto(linea)
            detalle = self._format_linea_detalle(linea)
            codigo_mp = (getattr(linea, "codigo_mp", "") or "").strip()
            codigo_sap_sugerido = self._format_linea_codigo_sap_sugerido(linea)
            cantidad_sap = self._format_linea_cantidad_sap(linea)
            precio_sap = self._format_linea_precio_sap(linea)

            extra_parts: list[str] = []
            if detalle:
                extra_parts.append(
                    "<div style=\"margin-top:6px;font-size:12px;line-height:1.5;color:#6b7280;\">"
                    f"{self._html_escape(detalle)}</div>"
                )
            if codigo_mp:
                extra_parts.append(
                    "<div style=\"margin-top:4px;font-size:11px;line-height:1.4;color:#64748b;\">"
                    f"Cod MP: {self._html_escape(codigo_mp)}</div>"
                )

            table_rows.append(
                "<tr>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;font-weight:700;white-space:nowrap;\">"
                f"{self._html_escape(getattr(linea, 'correlativo', '-') or '-')}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;\">"
                f"<div style=\"font-weight:700;color:#172033;\">{self._html_escape(producto)}</div>"
                f"{''.join(extra_parts)}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;font-weight:700;white-space:nowrap;\">"
                f"{self._html_escape(codigo_sap_sugerido)}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;white-space:nowrap;\">"
                f"{self._html_escape(self._format_quantity_with_unit(getattr(linea, 'cantidad', None), getattr(linea, 'unidad', '')))}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;white-space:nowrap;\">"
                f"{self._html_escape(cantidad_sap)}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;font-weight:700;white-space:nowrap;\">"
                f"{self._html_escape(self._format_currency(getattr(linea, 'precio_neto', 0)))}"
                "</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;font-weight:700;white-space:nowrap;\">"
                f"{self._html_escape(precio_sap)}"
                "</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;font-weight:700;white-space:nowrap;\">"
                f"{self._html_escape(self._format_currency(getattr(linea, 'total', 0)))}"
                "</td>"
                "</tr>"
            )

        desktop_table = (
            "<table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" "
            "style=\"border:1px solid #d8e1ee;border-radius:16px;overflow:hidden;border-collapse:separate;"
            "border-spacing:0;\">"
            "<tr>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Item</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Producto</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Codigo SAP sugerido</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Cantidad</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Cantidad SAP</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Precio neto</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Precio SAP</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Total</th>"
            "</tr>"
            f"{''.join(table_rows)}"
            "</table>"
        )
        mobile_cards = self._render_oc_lines_cards_html(lineas)
        return (
            f"<div class=\"desktop-only\">{desktop_table}</div>"
            f"<div class=\"mobile-only\" style=\"display:none;max-height:0;overflow:hidden;\">{mobile_cards}</div>"
        )

    def _render_oc_lines_cards_html(self, lineas: list) -> str:
        cards: list[str] = []
        for linea in lineas:
            producto = self._format_linea_producto(linea)
            codigo_mp = (getattr(linea, "codigo_mp", "") or "").strip()
            codigo_sap_sugerido = self._format_linea_codigo_sap_sugerido(linea)
            subtitle = f"Item {getattr(linea, 'correlativo', '-') or '-'}"
            if codigo_mp:
                subtitle = f"{subtitle} | Cod MP: {codigo_mp}"

            detail_rows = [
                ("Codigo SAP sugerido", codigo_sap_sugerido),
                ("Cantidad", self._format_quantity_with_unit(getattr(linea, "cantidad", None), getattr(linea, "unidad", ""))),
                ("Cantidad SAP", self._format_linea_cantidad_sap(linea)),
                ("Precio neto", self._format_currency(getattr(linea, "precio_neto", 0))),
                ("Precio SAP", self._format_linea_precio_sap(linea)),
                ("Total", self._format_currency(getattr(linea, "total", 0))),
            ]
            detalle = self._format_linea_detalle(linea)
            if detalle:
                detail_rows.append(("Detalle", detalle))

            cards.append(
                "<div class=\"mobile-summary-card\" "
                "style=\"margin:0 0 14px 0;border:1px solid #d8e1ee;border-radius:16px;background:#fbfdff;"
                "overflow:hidden;\">"
                "<div style=\"padding:14px 16px;background:#eef4ff;border-bottom:1px solid #d8e1ee;\">"
                f"<div class=\"mobile-summary-title\" style=\"font-size:16px;line-height:1.35;font-weight:700;color:#172033;\">"
                f"{self._html_escape(producto)}</div>"
                f"<div class=\"mobile-summary-meta\" style=\"margin-top:6px;font-size:13px;line-height:1.5;color:#4b5563;\">"
                f"{self._html_escape(subtitle)}</div>"
                "</div>"
                "<div style=\"padding:0 16px 8px 16px;\">"
                f"{self._render_key_value_table_html(detail_rows)}"
                "</div>"
                "</div>"
            )
        return "".join(cards)

    def _render_summary_table_html(self, rows: list[dict]) -> str:
        table_rows: list[str] = []
        for index, row in enumerate(rows):
            background = "#ffffff" if index % 2 == 0 else "#f8fbff"
            codigo_oc = str(row.get("codigo_oc") or "-").strip()
            codigo_cell = self._html_escape(codigo_oc)
            if codigo_oc and codigo_oc != "-":
                codigo_cell = (
                    f"<a href=\"{self._html_escape(self._portal_url(codigo_oc))}\" "
                    "style=\"color:#0f62fe;text-decoration:none;font-weight:700;\">"
                    f"{self._html_escape(codigo_oc)}</a>"
                )

            table_rows.append(
                "<tr>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;\">"
                f"{codigo_cell}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;\">"
                f"<strong>{self._html_escape(row.get('tipo_oc') or '-')}</strong></td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;\">"
                f"{self._html_escape(row.get('nombre_organismo') or '-')}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;font-weight:700;white-space:nowrap;\">"
                f"{self._html_escape(self._format_currency(row.get('total')))}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;white-space:nowrap;\">"
                f"{self._html_escape(self._format_datetime_short(row.get('created_at')))}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;\">"
                f"{self._render_badge_html(self._format_estado_sap(row.get('estado_interno', ''), row.get('fecha_ingreso', '')))}</td>"
                f"<td style=\"padding:12px 10px;border-bottom:1px solid #e5eaf3;background:{background};"
                "font-size:13px;line-height:1.5;color:#172033;\">"
                f"{self._render_badge_html(row.get('estado_mp') or '-')}</td>"
                "</tr>"
            )

        desktop_table = (
            "<table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" "
            "style=\"border:1px solid #d8e1ee;border-radius:16px;overflow:hidden;border-collapse:separate;"
            "border-spacing:0;\">"
            "<tr>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">OC</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Tipo</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Organismo</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Monto</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Llegada</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">SAP</th>"
            "<th align=\"left\" style=\"padding:12px 10px;background:#eaf2ff;border-bottom:1px solid #d8e1ee;"
            "font-size:11px;letter-spacing:0.12em;text-transform:uppercase;color:#50627a;\">Estado MP</th>"
            "</tr>"
            f"{''.join(table_rows)}"
            "</table>"
        )
        mobile_cards = self._render_summary_cards_html(rows)
        return (
            f"<div class=\"desktop-only\">{desktop_table}</div>"
            f"<div class=\"mobile-only\" style=\"display:none;max-height:0;overflow:hidden;\">{mobile_cards}</div>"
        )

    def _render_summary_cards_html(self, rows: list[dict]) -> str:
        cards: list[str] = []
        for row in rows:
            codigo_oc = str(row.get("codigo_oc") or "-").strip()
            codigo_label = self._html_escape(codigo_oc)
            if codigo_oc and codigo_oc != "-":
                codigo_label = (
                    f"<a href=\"{self._html_escape(self._portal_url(codigo_oc))}\" "
                    "style=\"color:#0f62fe;text-decoration:none;font-weight:700;word-break:break-word;\">"
                    f"{self._html_escape(codigo_oc)}</a>"
                )

            detalle_html = self._render_key_value_table_html([
                ("Monto", self._format_currency(row.get("total"))),
                ("Llegada", self._format_datetime_short(row.get("created_at"))),
                (
                    "SAP",
                    self._format_estado_sap(
                        row.get("estado_interno", ""),
                        row.get("fecha_ingreso", ""),
                    ),
                ),
                ("Estado MP", row.get("estado_mp") or "-"),
            ])

            cards.append(
                "<div class=\"mobile-summary-card\" "
                "style=\"margin:0 0 14px 0;border:1px solid #d8e1ee;border-radius:16px;background:#fbfdff;"
                "overflow:hidden;\">"
                "<div style=\"padding:14px 16px;background:#eef4ff;border-bottom:1px solid #d8e1ee;\">"
                f"<div class=\"mobile-summary-title\" style=\"font-size:16px;line-height:1.35;font-weight:700;color:#172033;\">"
                f"{codigo_label}</div>"
                f"<div class=\"mobile-summary-meta\" style=\"margin-top:6px;font-size:13px;line-height:1.5;color:#4b5563;\">"
                f"{self._html_escape(row.get('tipo_oc') or '-')} · {self._html_escape(row.get('nombre_organismo') or '-')}</div>"
                "</div>"
                "<div style=\"padding:0 16px 8px 16px;\">"
                f"{detalle_html}"
                "</div>"
                "</div>"
            )
        return "".join(cards)

    def _render_button_html(self, url: str, label: str) -> str:
        return (
            "<div style=\"margin-top:18px;\">"
            f"<a href=\"{self._html_escape(url)}\" "
            "style=\"display:inline-block;padding:12px 18px;border-radius:10px;background:#0f62fe;"
            "color:#ffffff;text-decoration:none;font-size:14px;font-weight:700;\">"
            f"{self._html_escape(label)}</a>"
            "</div>"
        )

    def _render_badge_html(self, value: str) -> str:
        text = (value or "-").strip() or "-"
        normalized = text.casefold()
        background = "#eef2f7"
        color = "#475569"

        if "ingresada" in normalized or "aceptada" in normalized:
            background = "#dcfce7"
            color = "#166534"
        elif "pend" in normalized or "nueva" in normalized or "revision" in normalized:
            background = "#fef3c7"
            color = "#92400e"
        elif "cancel" in normalized or "rechaz" in normalized or "error" in normalized:
            background = "#fee2e2"
            color = "#b91c1c"

        return (
            f"<span style=\"display:inline-block;padding:5px 9px;border-radius:999px;background:{background};"
            f"color:{color};font-size:11px;font-weight:700;line-height:1.3;\">"
            f"{self._html_escape(text)}</span>"
        )

    def _html_escape(self, value) -> str:
        return escape(str(value or ""))

    def _portal_url(self, codigo_oc: str) -> str:
        return (
            "https://www.mercadopublico.cl/PurchaseOrder/Modules/PO/"
            f"DetailsPurchaseOrder.aspx?codigoOC={codigo_oc}"
        )

    def _obtener_lineas_oc(self, codigo_oc: str) -> list:
        try:
            from app.repositories.oc_repository import get_lineas

            return get_lineas(codigo_oc)
        except Exception as e:
            logger.warning(
                f"No fue posible cargar lineas para OC {codigo_oc} al preparar notificacion email: {e}"
            )
            return []

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _destinatarios_cartera(self, cartera: str) -> list[str]:
        if not self._loaded:
            self.reload()
        return self._unique_emails(
            [
                dest.get("email", "").strip()
                for dest in self._vendedores.get((cartera or "").strip().upper(), [])
                if dest.get("email") and dest.get("activo")
            ]
        )

    def _build_cc_emails(self, cfg, to_emails: list[str]) -> list[str]:
        cc_email = (getattr(cfg, "notification_cc_email", "") or "").strip()
        if not cc_email:
            return []
        if cc_email.casefold() in {email.casefold() for email in to_emails}:
            return []
        return [cc_email]

    def _resolve_column_map(self, header_row: tuple) -> tuple[dict[str, Optional[int]], int]:
        normalized = [self._normalize_header(cell) for cell in header_row]
        mapped: dict[str, Optional[int]] = {"cartera": None, "nombre": None, "correo": None, "activo": None}

        for logical_name, aliases in HEADER_ALIASES.items():
            for idx, cell_name in enumerate(normalized):
                if cell_name in aliases:
                    mapped[logical_name] = idx
                    break

        has_header = mapped["cartera"] is not None or mapped["correo"] is not None
        if has_header:
            return mapped, 2

        return {
            "cartera": COL_CARTERA,
            "nombre": COL_NOMBRE,
            "correo": COL_CORREO,
            "activo": 3 if len(header_row) > 3 else None,
        }, 1

    def _get_cell(self, row: tuple, idx: Optional[int]):
        if idx is None or idx < 0 or idx >= len(row):
            return None
        return row[idx]

    def _parse_activo(self, value) -> bool:
        if value is None:
            return True
        normalized = str(value).strip().casefold()
        if not normalized:
            return True
        if normalized in ACTIVE_FALSE_VALUES:
            return False
        if normalized in ACTIVE_TRUE_VALUES:
            return True
        return True

    def _normalize_header(self, value) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        return text.replace(" ", "_")

    def _unique_emails(self, emails: list[str]) -> list[str]:
        unique: list[str] = []
        seen: set[str] = set()
        for email in emails:
            normalized = (email or "").strip()
            key = normalized.casefold()
            if not normalized or key in seen:
                continue
            seen.add(key)
            unique.append(normalized)
        return unique

    def _format_currency(self, amount) -> str:
        try:
            value = float(amount or 0)
        except (TypeError, ValueError):
            value = 0.0
        return f"${value:,.0f}".replace(",", ".")

    def _format_currency_smart(self, amount) -> str:
        try:
            value = float(amount)
        except (TypeError, ValueError):
            return "-"

        if value.is_integer():
            return f"${value:,.0f}".replace(",", ".")
        return (
            f"${value:,.4f}".rstrip("0").rstrip(".")
            .replace(",", "_")
            .replace(".", ",")
            .replace("_", ".")
        )

    def _format_quantity(self, amount) -> str:
        try:
            value = float(amount)
        except (TypeError, ValueError):
            return "-"

        if value.is_integer():
            return f"{int(value)}"
        return (
            f"{value:,.2f}"
            .replace(",", "_")
            .replace(".", ",")
            .replace("_", ".")
            .rstrip("0")
            .rstrip(",")
        )

    def _format_quantity_with_unit(self, amount, unit: str = "") -> str:
        qty = self._format_quantity(amount)
        unit_txt = (unit or "").strip()
        if qty == "-":
            return "-"
        return f"{qty} {unit_txt}".strip()

    def _format_linea_producto(self, linea) -> str:
        producto = (getattr(linea, "producto", "") or "").strip()
        if producto:
            return producto

        codigo_mp = (getattr(linea, "codigo_mp", "") or "").strip()
        if codigo_mp:
            return codigo_mp

        correlativo = getattr(linea, "correlativo", None)
        return f"Linea {correlativo}" if correlativo is not None else "Linea sin descripcion"

    def _format_linea_codigo_sap_sugerido(self, linea) -> str:
        codigo_sap = (getattr(linea, "itemcode_sap", "") or "").strip()
        return codigo_sap or "pendiente"

    def _format_linea_cantidad_sap(self, linea) -> str:
        cantidad = getattr(linea, "cantidad_sap", None)
        if cantidad is None:
            cantidad = getattr(linea, "cantidad", None)
        return self._format_quantity_with_unit(cantidad, getattr(linea, "unidad", ""))

    def _format_linea_precio_sap(self, linea) -> str:
        precio = getattr(linea, "precio_sap", None)
        if precio is None:
            precio = getattr(linea, "precio_neto", None)
        return self._format_currency_smart(precio)

    def _format_linea_detalle(self, linea) -> str:
        producto = self._format_linea_producto(linea).casefold()
        extras: list[str] = []

        for candidate in (
            getattr(linea, "especificacion_comprador", ""),
            getattr(linea, "especificacion_proveedor", ""),
        ):
            text = (candidate or "").strip()
            if not text:
                continue
            if text.casefold() == producto:
                continue
            if text.casefold() in {item.casefold() for item in extras}:
                continue
            extras.append(text)

        return " | ".join(extras[:2])

    def _format_date_only(self, value: str) -> str:
        raw = (value or "").strip()
        if not raw:
            return "-"
        try:
            return datetime.fromisoformat(raw).strftime("%d/%m/%Y")
        except ValueError:
            return raw[:10]

    def _format_datetime_short(self, value: str) -> str:
        raw = (value or "").strip()
        if not raw:
            return "-"
        try:
            return datetime.fromisoformat(raw).strftime("%H:%M")
        except ValueError:
            return raw[:16].replace("T", " ")

    def _format_estado_sap(self, estado_interno: str, fecha_ingreso: str = "") -> str:
        estado = (estado_interno or "Nueva").strip() or "Nueva"
        if estado.lower() != "ingresada":
            return f"No ingresada ({estado})"

        fecha_txt = self._format_datetime_short(fecha_ingreso)
        if fecha_txt != "-":
            return f"Ingresada ({fecha_txt})"
        return "Ingresada"
