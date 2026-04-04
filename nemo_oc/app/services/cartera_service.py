"""
Servicio de cartera de clientes.
Carga CARTERA(PBI).xlsx (hoja CARTERA 2026) y construye lookup en memoria.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from app.models.cartera import CarteraCliente
from app.db import get_connection
from app.utils.rut_utils import normalize_rut, normalize_rut_body

logger = logging.getLogger(__name__)

# Columnas de CARTERA(PBI).xlsx, hoja "CARTERA 2026" (índice base 0)
COL_RUT          = 0   # A: RUT
COL_RAZON        = 1   # B: RAZON
COL_COMUNA       = 2   # C: COMUNA
COL_REGION       = 3   # D: REGION (código numérico)
COL_VENDEDOR     = 4   # E: VENDEDOR
COL_COD_CLIENTE  = 5   # F: COD CLIENTE (CN...)
COL_INDUSTRIA    = 6   # G: INDUSTRIA
COL_SECTOR       = 7   # H: SECTOR
COL_CARTERA      = 8   # I: CARTERA
COL_REGION_NOMBRE = 9  # J: REGION NOMBRE

SHEET_NAME = "CARTERA 2026"


class CarteraService:
    """Gestiona la carga y búsqueda del catálogo de cartera de clientes."""

    def __init__(self):
        self._cache: Dict[str, CarteraCliente] = {}
        self._loaded = False

    def cargar_cartera_excel(self, path: str) -> Tuple[int, List[str]]:
        """Lee CARTERA(PBI).xlsx y guarda en SQLite + memoria."""
        try:
            import openpyxl
        except ImportError:
            return 0, ["openpyxl no está instalado."]

        errores: List[str] = []
        items: List[CarteraCliente] = []
        path_obj = Path(path)

        if not path_obj.exists():
            return 0, [f"Archivo no encontrado: {path}"]

        try:
            wb = openpyxl.load_workbook(str(path_obj), data_only=True, read_only=True)
            # Intentar hoja específica, fallback a activa
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            now = datetime.now().isoformat()
            filename = path_obj.name

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= COL_COD_CLIENTE:
                    continue
                raw_cod = row[COL_COD_CLIENTE]
                if not raw_cod:
                    continue

                cod_cliente = str(raw_cod).strip()
                if not cod_cliente:
                    continue

                items.append(CarteraCliente(
                    cod_cliente=cod_cliente,
                    rut=str(row[COL_RUT]).strip() if row[COL_RUT] is not None else "",
                    razon=str(row[COL_RAZON]).strip() if row[COL_RAZON] else "",
                    comuna=str(row[COL_COMUNA]).strip() if row[COL_COMUNA] else "",
                    region_cod=str(row[COL_REGION]).strip() if row[COL_REGION] is not None else "",
                    vendedor=str(row[COL_VENDEDOR]).strip() if row[COL_VENDEDOR] else "",
                    industria=str(row[COL_INDUSTRIA]).strip() if row[COL_INDUSTRIA] else "",
                    sector=str(row[COL_SECTOR]).strip() if row[COL_SECTOR] else "",
                    cartera=str(row[COL_CARTERA]).strip() if row[COL_CARTERA] else "",
                    region_nombre=str(row[COL_REGION_NOMBRE]).strip() if row[COL_REGION_NOMBRE] else "",
                    origen_archivo=filename,
                    created_at=now,
                    updated_at=now,
                ))
            wb.close()
        except Exception as e:
            return 0, [f"Error leyendo Excel: {e}"]

        if not items:
            return 0, ["No se encontraron registros válidos en el archivo."]

        count = self._guardar_cartera(items)
        self._reload_cache()
        logger.info(f"Cartera cargada: {count} registros desde {filename}")
        return count, errores

    def lookup(self, cod_cliente: str) -> Optional[CarteraCliente]:
        """Busca un cliente por COD CLIENTE. Retorna None si no existe."""
        if not self._loaded:
            self._reload_cache()
        return self._cache.get(str(cod_cliente).strip()) if cod_cliente else None

    def get_count(self) -> int:
        if not self._loaded:
            self._reload_cache()
        return len(self._cache)

    def search(self, query: str, limit: int = 8) -> List[CarteraCliente]:
        """Busca clientes por razon social, codigo cliente o RUT."""
        if not self._loaded:
            self._reload_cache()

        raw = (query or "").strip()
        if len(raw) < 2:
            return []

        q = raw.casefold()
        q_rut = normalize_rut(raw)
        q_rut_body = normalize_rut_body(raw)

        scored: List[Tuple[int, CarteraCliente]] = []
        seen: set[str] = set()

        for item in self._cache.values():
            if item.cod_cliente in seen:
                continue

            razon = (item.razon or "").casefold()
            cod_cliente = (item.cod_cliente or "").casefold()
            rut_norm = normalize_rut(item.rut)
            rut_body = normalize_rut_body(item.rut)

            score: Optional[int] = None

            if q_rut and rut_norm == q_rut:
                score = 0
            elif q_rut_body and rut_body == q_rut_body:
                score = 1
            elif cod_cliente.startswith(q):
                score = 2
            elif razon.startswith(q):
                score = 3
            elif q in cod_cliente:
                score = 4
            elif q in razon:
                score = 5
            elif q_rut_body and q_rut_body in rut_body:
                score = 6

            if score is None:
                continue

            scored.append((score, item))
            seen.add(item.cod_cliente)

        scored.sort(key=lambda entry: (entry[0], entry[1].razon.casefold(), entry[1].cod_cliente.casefold()))
        return [item for _, item in scored[:limit]]

    def reload(self):
        self._reload_cache()

    def _guardar_cartera(self, items: List[CarteraCliente]) -> int:
        conn = get_connection()
        count = 0
        try:
            now = datetime.now().isoformat()
            for item in items:
                conn.execute("""
                    INSERT INTO cartera_clientes
                        (cod_cliente, rut, razon, comuna, region_cod, vendedor,
                         industria, sector, cartera, region_nombre,
                         origen_archivo, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(cod_cliente) DO UPDATE SET
                        rut           = excluded.rut,
                        razon         = excluded.razon,
                        comuna        = excluded.comuna,
                        region_cod    = excluded.region_cod,
                        vendedor      = excluded.vendedor,
                        industria     = excluded.industria,
                        sector        = excluded.sector,
                        cartera       = excluded.cartera,
                        region_nombre = excluded.region_nombre,
                        origen_archivo= excluded.origen_archivo,
                        updated_at    = excluded.updated_at
                """, (
                    item.cod_cliente, item.rut, item.razon, item.comuna,
                    item.region_cod, item.vendedor, item.industria, item.sector,
                    item.cartera, item.region_nombre, item.origen_archivo,
                    item.created_at, now
                ))
                count += 1
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"Error guardando cartera: {e}")
            raise
        finally:
            conn.close()
        return count

    def _reload_cache(self):
        conn = get_connection()
        try:
            rows = conn.execute("SELECT * FROM cartera_clientes").fetchall()
            self._cache = {
                r["cod_cliente"]: CarteraCliente(
                    cod_cliente=r["cod_cliente"],
                    rut=r["rut"] or "",
                    razon=r["razon"] or "",
                    comuna=r["comuna"] or "",
                    region_cod=r["region_cod"] or "",
                    vendedor=r["vendedor"] or "",
                    industria=r["industria"] or "",
                    sector=r["sector"] or "",
                    cartera=r["cartera"] or "",
                    region_nombre=r["region_nombre"] or "",
                )
                for r in rows
            }
            self._loaded = True
        finally:
            conn.close()


# Singleton global
_service: Optional[CarteraService] = None


def get_cartera_service() -> CarteraService:
    global _service
    if _service is None:
        _service = CarteraService()
    return _service
