"""
Servicio para la Maestra de Materiales SAP.
Importa el archivo MAESTRA DE MATERIALES (PBI).xlsx y provee lookup
de codigo historico (viejo) → itemcode_sap (nuevo).
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from app.db import get_connection
from app.models.maestra_material import MaestraMaterial

logger = logging.getLogger(__name__)

_instance: Optional["MaestraService"] = None


def get_maestra_service() -> "MaestraService":
    global _instance
    if _instance is None:
        _instance = MaestraService()
    return _instance


class MaestraService:

    def __init__(self):
        self._cache: Dict[str, MaestraMaterial] = {}
        self._old_to_new: Dict[str, str] = {}
        self._reload_cache()

    def cargar_excel(self, path: str) -> Tuple[int, List[str]]:
        """
        Importa la hoja MATERIALES del archivo MAESTRA DE MATERIALES.
        Retorna (cantidad_importada, lista_errores).
        """
        errors: List[str] = []
        items: List[MaestraMaterial] = []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # Buscar hoja MATERIALES
            sheet_name = None
            for name in wb.sheetnames:
                if "MATERIALES" in name.upper() and "ANTIGUA" not in name.upper():
                    sheet_name = name
                    break
            if not sheet_name:
                sheet_name = wb.sheetnames[0]

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                return 0, ["Archivo vacío"]

            # Detectar columnas por header
            header = [str(c or "").strip() for c in rows[0]]
            col_map = {}
            for i, h in enumerate(header):
                h_lower = h.lower().replace("í", "i").replace("ó", "o").replace("ú", "u")
                if "numero" in h_lower and "articulo" in h_lower:
                    col_map["itemcode"] = i
                elif h_lower == "itemcode":
                    col_map["itemcode"] = i
                elif "descripci" in h_lower and "articulo" in h_lower:
                    col_map["descripcion"] = i
                elif "codigo" in h_lower and "histori" in h_lower:
                    col_map["historico"] = i
                elif "nombre" in h_lower and "grupo" in h_lower:
                    col_map["grupo"] = i
                elif "categor" in h_lower:
                    col_map["categoria"] = i
                elif "cant" in h_lower and "display" in h_lower:
                    col_map["cant_display"] = i
                elif "cant" in h_lower and "caja" in h_lower:
                    col_map["cant_caja"] = i

            if "itemcode" not in col_map:
                return 0, [f"No se encontró columna ItemCode en headers: {header[:5]}"]

            for row_idx, row in enumerate(rows[1:], start=2):
                try:
                    itemcode = str(row[col_map["itemcode"]] or "").strip()
                    if not itemcode:
                        continue

                    desc = str(row[col_map.get("descripcion", 1)] or "").strip() if "descripcion" in col_map else ""
                    hist = str(row[col_map.get("historico", 3)] or "").strip() if "historico" in col_map else ""
                    grupo = str(row[col_map.get("grupo", 2)] or "").strip() if "grupo" in col_map else ""
                    cat = str(row[col_map.get("categoria", 4)] or "").strip() if "categoria" in col_map else ""

                    def _float(val):
                        try:
                            return float(val) if val else 0.0
                        except (ValueError, TypeError):
                            return 0.0

                    cant_d = _float(row[col_map["cant_display"]]) if "cant_display" in col_map else 0.0
                    cant_c = _float(row[col_map["cant_caja"]]) if "cant_caja" in col_map else 0.0

                    items.append(MaestraMaterial(
                        itemcode_sap=itemcode,
                        descripcion=desc,
                        codigo_historico=hist or itemcode,
                        grupo=grupo,
                        categoria=cat,
                        cant_display=cant_d,
                        cant_caja_master=cant_c,
                    ))
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {e}")

        except Exception as e:
            return 0, [f"Error leyendo archivo: {e}"]

        if not items:
            return 0, errors or ["No se encontraron materiales"]

        # Guardar en BD
        count = self._guardar(items, Path(path).name)
        self._reload_cache()
        logger.info(f"Maestra importada: {count} materiales desde {Path(path).name}")
        return count, errors

    def lookup(self, itemcode_sap: str) -> Optional[MaestraMaterial]:
        return self._cache.get(itemcode_sap)

    def lookup_by_old_code(self, old_code: str) -> Optional[MaestraMaterial]:
        """Busca material por codigo historico (viejo)."""
        new_code = self._old_to_new.get(old_code)
        if new_code:
            return self._cache.get(new_code)
        # Fallback: tal vez old_code ya es el nuevo
        return self._cache.get(old_code)

    def count(self) -> int:
        return len(self._cache)

    def reload(self):
        self._reload_cache()

    def _guardar(self, items: List[MaestraMaterial], origen: str) -> int:
        conn = get_connection()
        now = datetime.now().isoformat()
        count = 0
        try:
            for item in items:
                conn.execute("""
                    INSERT OR REPLACE INTO maestra_materiales
                    (itemcode_sap, descripcion, codigo_historico, grupo, categoria,
                     cant_display, cant_caja_master, origen_archivo, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    item.itemcode_sap, item.descripcion, item.codigo_historico,
                    item.grupo, item.categoria, item.cant_display, item.cant_caja_master,
                    origen, now, now
                ))
                count += 1
            conn.commit()
        finally:
            conn.close()
        return count

    def _reload_cache(self):
        self._cache.clear()
        self._old_to_new.clear()
        conn = get_connection()
        try:
            rows = conn.execute("SELECT * FROM maestra_materiales").fetchall()
            for r in rows:
                item = MaestraMaterial(
                    itemcode_sap=r["itemcode_sap"],
                    descripcion=r["descripcion"] or "",
                    codigo_historico=r["codigo_historico"] or "",
                    grupo=r["grupo"] or "",
                    categoria=r["categoria"] or "",
                    cant_display=r["cant_display"] or 0.0,
                    cant_caja_master=r["cant_caja_master"] or 0.0,
                    origen_archivo=r["origen_archivo"] or "",
                )
                self._cache[item.itemcode_sap] = item
                if item.codigo_historico and item.codigo_historico != item.itemcode_sap:
                    self._old_to_new[item.codigo_historico] = item.itemcode_sap
        except Exception:
            pass
        finally:
            conn.close()
