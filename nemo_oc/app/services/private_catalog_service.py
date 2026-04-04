"""
Importacion y consulta de catalogos privados por holding.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from app.db import get_connection

logger = logging.getLogger(__name__)


@dataclass
class PrivateHoldingInfo:
    id: str
    nombre: str
    prefijo: str
    parser_type: str
    homo_file: str = ""
    catalog_count: int = 0


_HEADER_ALIASES = {
    "codigo_cliente": {
        "codigo material",
        "codigo",
        "codigo cliente",
        "cod cliente",
        "cod",
        "codigo antiguo",
        "codigo interno",
        "codigo item",
    },
    "descripcion": {
        "descripcion",
        "descripcion sap",
        "detalle",
        "producto",
        "glosa",
        "articulo",
        "nombre producto",
    },
    "itemcode_sap": {
        "codigo nemo",
        "cod nemo",
        "itemcode",
        "itemcode sap",
        "cod sap",
        "codigo sap",
        "item sap",
    },
    "precio_ref": {
        "precio",
        "precio ref",
        "precio referencia",
        "precio unitario",
        "valor",
    },
}


def list_private_holdings() -> list[PrivateHoldingInfo]:
    conn = get_connection()
    try:
        rows = conn.execute(
            """
            SELECT
                h.id,
                h.nombre,
                h.prefijo,
                h.parser_type,
                h.homo_file,
                COUNT(p.codigo_cliente) AS catalog_count
            FROM holdings h
            LEFT JOIN homologacion_privados p
                ON p.holding_id = h.id
            WHERE h.activo = 1
            GROUP BY h.id, h.nombre, h.prefijo, h.parser_type, h.homo_file
            ORDER BY h.nombre
            """
        ).fetchall()
        return [
            PrivateHoldingInfo(
                id=row["id"],
                nombre=row["nombre"],
                prefijo=row["prefijo"],
                parser_type=row["parser_type"],
                homo_file=row["homo_file"] or "",
                catalog_count=row["catalog_count"] or 0,
            )
            for row in rows
        ]
    finally:
        conn.close()


def import_private_catalog(holding_id: str, path: str, original_filename: Optional[str] = None) -> tuple[int, list[str]]:
    try:
        import openpyxl
    except ImportError:
        return 0, ["openpyxl no esta instalado."]

    path_obj = Path(path)
    if not path_obj.exists():
        return 0, [f"Archivo no encontrado: {path}"]

    holding = _get_holding(holding_id)
    if not holding:
        return 0, [f"Holding no existe: {holding_id}"]

    try:
        wb = openpyxl.load_workbook(str(path_obj), data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return 0, [f"Error leyendo Excel: {e}"]

    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return 0, ["El archivo esta vacio."]
        header_map = _resolve_headers(header_row)
        if "codigo_cliente" not in header_map or "itemcode_sap" not in header_map:
            return 0, ["No se encontraron columnas obligatorias de codigo e itemcode."]

        items: list[tuple[str, str, str, float]] = []
        errors: list[str] = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            codigo = _cell_str(row, header_map.get("codigo_cliente"))
            itemcode = _cell_str(row, header_map.get("itemcode_sap"))
            if not codigo and not itemcode:
                continue
            if not codigo:
                errors.append(f"Fila {idx}: codigo vacio")
                continue
            if not itemcode:
                errors.append(f"Fila {idx}: itemcode vacio")
                continue

            descripcion = _cell_str(row, header_map.get("descripcion"))
            precio_ref = _cell_float(row, header_map.get("precio_ref"))
            items.append((codigo, descripcion, itemcode, precio_ref))

        if not items:
            return 0, errors or ["No se encontraron registros validos."]

        imported = _save_private_catalog(
            holding_id=holding_id,
            items=items,
            filename=path_obj.name,
            original_filename=original_filename or path_obj.name,
            file_path=str(path_obj),
        )
        return imported, errors
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _resolve_headers(header_row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        header = _normalize_header(raw)
        if not header:
            continue
        for key, aliases in _HEADER_ALIASES.items():
            if header in aliases and key not in mapping:
                mapping[key] = idx
    return mapping


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    text = " ".join(text.split())
    return text


def _cell_str(row, index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return str(value).strip() if value is not None else ""


def _cell_float(row, index: Optional[int]) -> float:
    if index is None or index >= len(row):
        return 0.0
    value = row[index]
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        try:
            text = str(value).strip().replace(".", "").replace(",", ".")
            return float(text)
        except Exception:
            return 0.0


def _get_holding(holding_id: str) -> Optional[PrivateHoldingInfo]:
    holdings = {item.id: item for item in list_private_holdings()}
    return holdings.get(holding_id)


def _save_private_catalog(
    holding_id: str,
    items: list[tuple[str, str, str, float]],
    filename: str,
    original_filename: str,
    file_path: str,
) -> int:
    now = datetime.now().isoformat()
    conn = get_connection()
    try:
        for codigo, descripcion, itemcode, precio_ref in items:
            conn.execute(
                """
                INSERT INTO homologacion_privados
                    (codigo_cliente, holding_id, descripcion, itemcode_sap, precio_ref, origen_archivo, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(codigo_cliente, holding_id) DO UPDATE SET
                    descripcion = excluded.descripcion,
                    itemcode_sap = excluded.itemcode_sap,
                    precio_ref = excluded.precio_ref,
                    origen_archivo = excluded.origen_archivo,
                    updated_at = excluded.updated_at
                """,
                (codigo, holding_id, descripcion, itemcode, precio_ref, filename, now, now),
            )

        conn.execute(
            "UPDATE holding_catalog_files SET activo = 0, updated_at = ? WHERE holding_id = ? AND catalog_kind = 'homologacion'",
            (now, holding_id),
        )
        conn.execute(
            """
            INSERT INTO holding_catalog_files
                (holding_id, catalog_kind, filename, original_filename, file_path, activo, created_at, updated_at)
            VALUES (?, 'homologacion', ?, ?, ?, 1, ?, ?)
            """,
            (holding_id, filename, original_filename, file_path, now, now),
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"Error guardando catalogo privado {holding_id}: {e}")
        raise
    finally:
        conn.close()
    return len(items)
