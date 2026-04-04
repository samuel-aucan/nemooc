"""
Servicio para importar y consultar licitaciones históricas (lic.xlsx).

CAMBIOS RESPECTO AL MOTOR ANTERIOR:
  - Fuente: lic.xlsx en lugar de Licsccc.xlsx
  - 'codigo final' se usa directamente como itemcode_sap (no hay mapeo via maestra)
  - Scoring: Jaccard ponderado bidireccional (reemplaza matched/total_tokens)
  - Word-safe: compara conjuntos de palabras completas (no substring)
  - Umbral mínimo: score < _MIN_SCORE → no se muestra sugerencia
  - Pesos por tipo de token: números+unidad > nombres largos > palabras cortas
"""
import logging
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from app.db import get_connection
from app.models.licitacion_ref import LicitacionRef, SugerenciaProducto
from app.repositories import licitaciones_repo
from app.utils.rut_utils import rut_to_cliente_sap

logger = logging.getLogger(__name__)

_instance: Optional["LicitacionesService"] = None

# ── Constantes de scoring ───────────────────────────────────────────────────

# Umbral mínimo: por debajo de este score no se muestra ninguna sugerencia.
# 0.28 significa que la intersección ponderada debe ser al menos ~28% de la unión.
_MIN_SCORE: float = 0.28

# Cuántos candidatos traer de BD para rescoring en Python.
# Más alto = mejor recall, más trabajo de Python (sigue siendo rápido con 80).
_CANDIDATES_LIMIT: int = 80

# Stopwords: palabras que no aportan información al matching
_STOPWORDS = frozenset((
    "de del la las el los un una uno para por con sin que como mas en al"
    " und unidad unidades caja cajas sobre paquete pqte set kit par"
    " tipo uso marca modelo ref referencia numero num nro cod codigo"
    " segun especificacion aprox aproximado similar equivalente"
    " descripcion producto articulo item material und cada"
).split())

# Número + unidad: "500ml", "1000cc", "250mg", "10mm", "5fr", "1000ui" etc.
_UNIT_RE = re.compile(
    r'^\d+[\.,]?\d*\s*(ml|cc|mg|mcg|ug|gr|g|kg|lt|l|mm|cm|m|un|pcs|pzs|fr|f|ui|iu|x)$',
    re.IGNORECASE,
)

# Número puro: "500", "1000", "0.9"
_NUM_RE = re.compile(r'^\d+[\.,]?\d*$')


# ── Funciones públicas ──────────────────────────────────────────────────────

def get_licitaciones_service() -> "LicitacionesService":
    global _instance
    if _instance is None:
        _instance = LicitacionesService()
    return _instance


# ── Helpers de texto ────────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    """Lowercase, sin acentos, whitespace compacto."""
    nfkd = unicodedata.normalize("NFKD", text)
    sin_acentos = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", sin_acentos.lower().strip())


def _extract_tokens(text: str) -> List[str]:
    """
    Extrae tokens significativos: palabras >= 3 chars, no stopwords.
    Usa findall([a-z0-9]+) sobre texto normalizado → solo palabras completas,
    sin riesgo de substring matching.
    """
    normalized = _normalize(text)
    words = re.findall(r"[a-z0-9]+", normalized)
    return [w for w in words if len(w) >= 3 and w not in _STOPWORDS]


def _token_weight(token: str) -> float:
    """
    Peso por tipo de token. Tokens más específicos pesan más.

    Ejemplos:
      "500ml"    → 2.5  (número + unidad: muy específico)
      "1000"     → 2.0  (número puro: específico)
      "fentanilo"→ 1.5  (8+ chars: nombre técnico)
      "nitrilo"  → 1.5  (8+ chars)
      "suero"    → 1.0  (normal)
      "tal"      → 0.7  (3 chars: ambiguo, puede matchear en muchas palabras)
    """
    if _UNIT_RE.match(token):
        return 2.5
    if _NUM_RE.match(token):
        return 2.0
    if len(token) >= 8:
        return 1.5
    if len(token) <= 3:
        return 0.7
    return 1.0


def _jaccard_weighted(query_tokens: List[str], hist_tokens: List[str]) -> float:
    """
    Jaccard ponderado bidireccional.

    score = suma_pesos(intersección) / suma_pesos(unión)

    Ventajas vs el score anterior (matched / len(query_tokens)):
      - Bidireccional: penaliza si la descripción histórica tiene muchas palabras
        que no aparecen en la query (evita matches vagos).
      - Ponderado: tokens específicos (números, unidades, nombres largos) pesan
        más que tokens cortos o genéricos.
      - Word-safe: opera sobre sets de palabras completas, no substrings.

    Ejemplo:
      query:   {"guante", "nitrilo"}
      histor.: {"guante", "nitrilo", "desechable", "talla", "azul"}
      intersec = {guante, nitrilo}  → pesos: 1.0 + 1.5 = 2.5
      union    = 5 tokens           → pesos: 2.5 + 1.0 + 0.7 + 0.7 = 4.9
      score    = 2.5 / 4.9 = 0.51  ✓ razonable

      Si fuera: query={"guante"}, histor.={"guante", "ortopedico", "quirurgico",...}
      intersec → peso 1.0
      union    → peso ~6.0
      score    = 1.0 / 6.0 = 0.17  → se descarta con umbral 0.28  ✓
    """
    q_set = set(query_tokens)
    h_set = set(hist_tokens)

    intersection = q_set & h_set
    union = q_set | h_set

    if not union:
        return 0.0

    inter_w = sum(_token_weight(t) for t in intersection)
    union_w = sum(_token_weight(t) for t in union)

    return inter_w / union_w if union_w > 0 else 0.0


# ── Servicio ────────────────────────────────────────────────────────────────

class LicitacionesService:

    def __init__(self):
        self._count = 0
        self._reload_count()

    # ── Importación ────────────────────────────────────────────────────────

    def importar_lic(self, path: str) -> Tuple[int, List[str]]:
        """
        Importa lic.xlsx (Sheet1).

        Diferencia clave vs el importador anterior (Licsccc.xlsx):
          - Lee columna 'codigo final' directamente como itemcode_sap.
          - NO necesita lookup en MaestraService para resolver código viejo → nuevo.
          - Deduplicación: agrupa (descripcion_norm, codigo_final), cuenta frecuencia.

        Retorna (cantidad_refs_guardadas, lista_de_errores).
        """
        errors: List[str] = []

        try:
            import openpyxl
            logger.info(f"Leyendo {Path(path).name} ...")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # Sheet1 (índice 0) tiene los datos de licitaciones
            ws = wb.worksheets[0]

            # Detectar columnas por header (tolerante a mayúsculas/acentos)
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [str(c or "").strip() for c in header_row]

            col_idx: Dict[str, int] = {}
            for i, h in enumerate(header):
                h_norm = _normalize(h)
                if h == "Description":
                    col_idx["desc"] = i
                elif h == "ProductDesc":
                    col_idx["pdesc"] = i
                elif h == "CompetitorName":
                    col_idx["comp"] = i
                elif h == "Estado Gestion":
                    col_idx["estado"] = i
                elif h_norm in ("codigo final", "codigo_final"):
                    col_idx["codigo_final"] = i
                elif h_norm == "rut":
                    col_idx["rut"] = i

            required = {"desc", "comp", "estado", "codigo_final"}
            missing = required - set(col_idx.keys())
            if missing:
                wb.close()
                return 0, [
                    f"Columnas faltantes en Sheet1: {missing}. "
                    f"Headers encontrados: {header[:15]}"
                ]

            # Leer filas: filtrar solo Nemo Chile S.A. + Cotizada
            pairs: Dict[tuple, dict] = defaultdict(
                lambda: {"freq": 0, "desc_orig": "", "pdesc": ""}
            )
            nemo_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                comp   = str(row[col_idx["comp"]]   or "").strip()
                estado = str(row[col_idx["estado"]]  or "").strip()

                if comp != "NEMO CHILE S.A." or estado != "Cotizada":
                    continue

                nemo_count += 1
                desc         = str(row[col_idx["desc"]]         or "").strip()
                codigo_final = str(row[col_idx["codigo_final"]] or "").strip()
                pdesc = (
                    str(row[col_idx["pdesc"]] or "").strip()
                    if "pdesc" in col_idx else ""
                )

                # Descartar filas sin código o sin descripción
                if not codigo_final or not desc:
                    continue

                raw_rut = str(row[col_idx["rut"]] or "").strip() if "rut" in col_idx else ""
                cliente_sap = rut_to_cliente_sap(raw_rut) if raw_rut else ""

                desc_norm = _normalize(desc)
                key = (desc_norm, cliente_sap, codigo_final)
                pairs[key]["freq"] += 1
                if not pairs[key]["desc_orig"]:
                    pairs[key]["desc_orig"] = desc
                if not pairs[key]["pdesc"] and pdesc:
                    pairs[key]["pdesc"] = pdesc

            wb.close()
            logger.info(
                f"lic.xlsx: {nemo_count} filas Nemo+Cotizada → {len(pairs)} pares únicos"
            )

        except Exception as e:
            return 0, [f"Error leyendo archivo: {e}"]

        if not pairs:
            return 0, errors or ["No se encontraron datos Nemo+Cotizada en Sheet1"]

        # Construir LicitacionRef — código final ya ES el itemcode_sap
        refs: List[LicitacionRef] = []
        for (desc_norm, cliente_sap, codigo_final), info in pairs.items():
            refs.append(LicitacionRef(
                rut_comprador=cliente_sap,
                descripcion_comprador=info["desc_orig"],
                descripcion_norm=desc_norm,
                producto_code_old=codigo_final,  # guardamos para trazabilidad
                itemcode_sap=codigo_final,        # directo, sin mapeo
                descripcion_nemo=info["pdesc"],
                frecuencia=info["freq"],
                origen_archivo=Path(path).name,
            ))

        count = self._guardar(refs, Path(path).name)
        self._reload_count()

        logger.info(f"Licitaciones guardadas: {count} refs desde {Path(path).name}")
        return count, errors

    def importar_licsccc(self, path: str) -> Tuple[int, List[str]]:
        """Alias de importar_lic — mantiene compatibilidad con config_frame."""
        return self.importar_lic(path)

    # ── Búsqueda / Scoring ─────────────────────────────────────────────────

    def buscar_sugerencias(self, texto: str, rut_oc: str = None, max_results: int = 5) -> List[SugerenciaProducto]:
        """
        Busca sugerencias para una descripción de OC no-CM.

        Flujo Jerárquico:
          1. Exact Match + RUT (Score 1.0 directo)
          2. Jaccard + RUT (Bono de score)
          3. Jaccard Global (Histórico licitaciones general)
          4. Maestra SAP Fallback (solo si no hay matches en historial)
        """
        from app.repositories import maestra_repo
        from app.utils.rut_utils import rut_to_cliente_sap

        query_tokens = _extract_tokens(texto)
        if not query_tokens:
            return []

        desc_norm = _normalize(texto)
        cliente_sap = rut_to_cliente_sap(rut_oc) if rut_oc else None

        # FASE 1: Exact Match por RUT (Bala de plata)
        if cliente_sap:
            exact_rows = licitaciones_repo.get_exact_candidates(desc_norm, cliente_sap)
            if exact_rows:
                mejor = exact_rows[0]
                return [SugerenciaProducto(
                    itemcode_sap=mejor["itemcode_sap"],
                    descripcion_sap=mejor.get("descripcion_nemo") or "",
                    descripcion_match=mejor.get("descripcion_comprador") or "",
                    frecuencia=mejor.get("frecuencia", 1),
                    score=1.0,
                )]

        best_by_code: Dict[str, SugerenciaProducto] = {}

        # Función auxiliar para rescore
        def _score_candidates(rows, boost=0.0, is_maestra=False):
            for r in rows:
                itemcode = r.get("itemcode_sap")
                if not itemcode: continue

                if is_maestra:
                    hist_tokens = _extract_tokens(r.get("descripcion_sap") or "")
                else:
                    hist_desc = r.get("descripcion_norm") or _normalize(r.get("descripcion_comprador") or "")
                    hist_tokens = _extract_tokens(hist_desc)

                if not hist_tokens: continue

                score = _jaccard_weighted(query_tokens, hist_tokens) + boost
                
                umbral = 0.20 if is_maestra else _MIN_SCORE
                if score < umbral:
                    continue

                score = min(score, 0.99)  # Max 0.99 para no pisar el 1.0 exacto

                existing = best_by_code.get(itemcode)
                if existing is None or score > existing.score:
                    if is_maestra:
                        best_by_code[itemcode] = SugerenciaProducto(
                            itemcode_sap=itemcode,
                            descripcion_sap=r.get("descripcion_sap") or "",
                            descripcion_match="[BD MAESTRA SAP]",
                            frecuencia=1,
                            score=score,
                        )
                    else:
                        best_by_code[itemcode] = SugerenciaProducto(
                            itemcode_sap=itemcode,
                            descripcion_sap=r.get("descripcion_nemo") or "",
                            descripcion_match=r.get("descripcion_comprador") or "",
                            frecuencia=r.get("frecuencia", 1),
                            score=score,
                        )

        # FASE 2: Similitud misma clínica/hospital (RUT)
        if cliente_sap:
            rows_f2 = licitaciones_repo.get_candidates(query_tokens, limit=_CANDIDATES_LIMIT, rut=cliente_sap)
            _score_candidates(rows_f2, boost=0.1)

        # FASE 3: Similitud global (cualquier RUT)
        rows_f3 = licitaciones_repo.get_candidates(query_tokens, limit=_CANDIDATES_LIMIT, rut=None)
        _score_candidates(rows_f3, boost=0.0)

        results = sorted(best_by_code.values(), key=lambda x: (-x.score, -x.frecuencia))

        # FASE 4: Backup usando diccionarios de la Maestra SAP
        if not results:
            rows_m = maestra_repo.search_by_keywords(query_tokens, limit=15)
            # Aplicamos una penalización al buscar directo en maestra para no dar falsos positivos tan altos
            _score_candidates(rows_m, boost=-0.1, is_maestra=True)
            results = sorted(best_by_code.values(), key=lambda x: (-x.score, -x.frecuencia))

        return results[:max_results]

    # ── Conteo ─────────────────────────────────────────────────────────────

    def count(self) -> int:
        return self._count

    def reload(self):
        self._reload_count()

    # ── Interno ────────────────────────────────────────────────────────────

    def _guardar(self, refs: List[LicitacionRef], origen: str) -> int:
        conn = get_connection()
        now = datetime.now().isoformat()
        try:
            conn.execute("DELETE FROM licitaciones_ref")
            count = 0
            for ref in refs:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO licitaciones_ref
                    (rut_comprador, descripcion_comprador, descripcion_norm, producto_code_old,
                     itemcode_sap, descripcion_nemo, frecuencia,
                     origen_archivo, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ref.rut_comprador,
                        ref.descripcion_comprador,
                        ref.descripcion_norm,
                        ref.producto_code_old,
                        ref.itemcode_sap or None,
                        ref.descripcion_nemo,
                        ref.frecuencia,
                        origen,
                        now,
                        now,
                    ),
                )
                count += 1
            conn.commit()
            return count
        finally:
            conn.close()

    def _reload_count(self):
        try:
            self._count = licitaciones_repo.count_licitaciones()
        except Exception:
            self._count = 0
